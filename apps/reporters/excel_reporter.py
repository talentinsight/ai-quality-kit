"""Excel report generation for orchestrator results."""

import json
from pathlib import Path
from typing import Dict, List, Any, cast
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter


def write_excel(path: str, data: Dict[str, Any]) -> None:
    """Write comprehensive Excel report from JSON data.
    
    Args:
        path: Output Excel file path
        data: JSON report data from build_json()
    """
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create sheets in order
    _create_summary_sheet(wb, data)
    _create_detailed_sheet(wb, data)
    _create_api_details_sheet(wb, data)
    _create_inputs_expected_sheet(wb, data)
    
    # Optional sheets for red team results
    if data.get("adversarial_details") and len(data["adversarial_details"]) > 0:
        _create_adversarial_sheet(wb, data)
    
    if data.get("coverage") and data["coverage"]:
        _create_coverage_sheet(wb, data)
    
    # Optional sheet for resilience results
    if data.get("resilience") and data["resilience"].get("details"):
        _create_resilience_details_sheet(wb, data)
    
    # Optional sheets for new smoke suites
    if data.get("compliance_smoke") and data["compliance_smoke"].get("details"):
        _create_compliance_details_sheet(wb, data)
    
    if data.get("bias_smoke") and data["bias_smoke"].get("details"):
        _create_bias_details_sheet(wb, data)
    
    # Optional sheet for guardrails results
    if data.get("guardrails") and data["guardrails"].get("details"):
        _create_guardrails_details_sheet(wb, data)
    
    # Optional sheet for logs
    if data.get("logs") and data["logs"].get("entries"):
        _create_logs_sheet(wb, data)
    
    # Optional sheets for RAG Reliability & Robustness (Prompt Robustness)
    if data.get("rag_reliability_robustness") and data["rag_reliability_robustness"].get("prompt_robustness"):
        prompt_robustness_data = data["rag_reliability_robustness"]["prompt_robustness"]
        
        # Import structure sheet writers
        from apps.reporting.structure_sheet import (
            write_structure_sheet,
            write_structure_prompts_sheet,
            write_structure_diffs_sheet
        )
        
        # Write structure sheets if data is available
        if prompt_robustness_data.get("structure_rows"):
            write_structure_sheet(wb, prompt_robustness_data["structure_rows"])
        
        if prompt_robustness_data.get("prompts_rows"):
            write_structure_prompts_sheet(wb, prompt_robustness_data["prompts_rows"])
        
        if prompt_robustness_data.get("diffs_rows"):
            write_structure_diffs_sheet(wb, prompt_robustness_data["diffs_rows"])
    
    # Optional sheet for Compare Mode results
    if data.get("compare") and data["compare"].get("enabled"):
        _create_compare_sheet(wb, data)
    
    # Save workbook
    wb.save(path)


def _create_summary_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Summary sheet with run overview."""
    ws = cast(Worksheet, wb.create_sheet("Summary", 0))
    
    headers = [
        "run_id", "started_at", "finished_at", "target_mode", "ground_truth", 
        "suites", "total_tests", "pass_rate", "faithfulness_avg", "context_recall_avg", 
        "attack_success_rate", "warm_p95_ms", "provider", "model", "gate", "elapsed_ms",
        "retrieval_top_k", "retrieval_note", "profile", "concurrency"
    ]
    
    # Add MCP-specific headers if using MCP mode
    run_meta = data.get("run", {})
    if run_meta.get("client_kind") == "mcp":
        headers.extend([
            "client_kind", "endpoint_host", "tool_name", "arg_shape", 
            "output_jsonpath", "contexts_jsonpath"
        ])
    
    # Add resilience headers if resilience data exists
    if data.get("resilience") and data["resilience"].get("summary"):
        headers.extend([
            "resilience_samples", "resilience_success_rate", "resilience_timeouts",
            "resilience_5xx", "resilience_429", "resilience_circuit_open", "resilience_p50_ms", "resilience_p95_ms"
        ])
    
    # Add compliance_smoke headers if data exists
    if data.get("compliance_smoke") and data["compliance_smoke"].get("summary"):
        headers.extend([
            "compliance_cases_scanned", "compliance_pii_hits", "compliance_rbac_checks", 
            "compliance_rbac_violations", "compliance_pass"
        ])
    
    # Add bias_smoke headers if data exists
    if data.get("bias_smoke") and data["bias_smoke"].get("summary"):
        headers.extend([
            "bias_pairs", "bias_metric", "bias_fails", "bias_fail_ratio", "bias_pass"
        ])
    
    # Add guardrails headers if data exists
    if data.get("guardrails") and data["guardrails"].get("summary"):
        headers.extend([
            "guardrails_total", "guardrails_pass", "guardrails_fail", "guardrails_warn", "guardrails_skip"
        ])
    
    # Add dataset metadata headers (additive)
    summary = data.get("summary", {})
    
    # Add Ragas headers if RAG quality suite has Ragas metrics
    rag_summary = summary.get("rag_quality", {})
    if rag_summary.get("ragas"):
        headers.extend([
            "ragas_faithfulness", "ragas_answer_relevancy", 
            "ragas_context_precision", "ragas_context_recall", "ragas_thresholds_passed"
        ])
    
    # Add Promptfoo headers if Promptfoo suite exists
    promptfoo_summary = summary.get("promptfoo", {})
    if promptfoo_summary:
        headers.extend([
            "promptfoo_total", "promptfoo_passed", "promptfoo_pass_rate"
        ])
    
    # Add MCP Security headers if MCP Security suite exists
    mcp_summary = summary.get("mcp_security", {})
    
    # Add RAG Reliability & Robustness headers if prompt robustness data exists
    rag_reliability_summary = data.get("rag_reliability_robustness", {}).get("prompt_robustness", {}).get("rollups", {})
    if rag_reliability_summary:
        headers.extend([
            "prompt_robustness_capacity_lift", "prompt_robustness_stability_simple", 
            "prompt_robustness_stability_cot", "prompt_robustness_stability_scaffold",
            "prompt_robustness_contract_adherence"
        ])
    
    # Add Embedding Robustness headers if data exists
    embedding_robustness_data = data.get("rag_reliability_robustness", {}).get("embedding_robustness")
    if embedding_robustness_data:
        headers.extend([
            "embedding_robustness_recall_at_k", "embedding_robustness_overlap_at_k",
            "embedding_robustness_answer_stability", "embedding_robustness_low_agreement_flag",
            "embedding_robustness_fallback_triggered", "embedding_robustness_hybrid_gain_delta",
            "embedding_robustness_paraphrase_count"
        ])
    if mcp_summary:
        headers.extend([
            "mcp_security_tests", "mcp_security_passed", "mcp_p95_latency_ms", 
            "mcp_slo_met", "mcp_schema_stable", "mcp_scope_denied"
        ])
    if summary.get("dataset_source"):
        headers.extend([
            "dataset_source", "dataset_version", "estimated_tests"
        ])
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)  # type: ignore
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write data row
    run_meta = data.get("run", {})
    summary = data.get("summary", {})
    
    row_data = [
        run_meta.get("run_id", ""),
        run_meta.get("started_at", ""),
        run_meta.get("finished_at", ""),
        run_meta.get("target_mode", ""),
        run_meta.get("ground_truth", ""),
        ",".join(run_meta.get("suites", [])),
        summary.get("overall", {}).get("total_tests", 0),
        summary.get("overall", {}).get("pass_rate", 0.0),
        summary.get("rag_quality", {}).get("avg_faithfulness", 0.0),
        summary.get("rag_quality", {}).get("avg_context_recall", 0.0),
        summary.get("safety", {}).get("attack_success_rate", 0.0),
        summary.get("performance", {}).get("p95_latency_ms", 0),
        run_meta.get("provider", ""),
        run_meta.get("model", ""),
        run_meta.get("gate", ""),
        run_meta.get("elapsed_ms", 0),
        run_meta.get("retrieval_top_k", ""),
        run_meta.get("retrieval_note", ""),
        run_meta.get("profile", ""),
        run_meta.get("concurrency", "")
    ]
    
    # Add MCP-specific data if using MCP mode
    if run_meta.get("client_kind") == "mcp":
        extraction_config = run_meta.get("extraction_config", {})
        row_data.extend([
            run_meta.get("client_kind", ""),
            run_meta.get("endpoint_host", ""),
            run_meta.get("tool_name", ""),
            run_meta.get("arg_shape", ""),
            extraction_config.get("output_jsonpath", ""),
            extraction_config.get("contexts_jsonpath", "")
        ])
    
    # Add resilience data if available
    if data.get("resilience") and data["resilience"].get("summary"):
        resilience_summary = data["resilience"]["summary"]
        row_data.extend([
            resilience_summary.get("samples", 0),
            resilience_summary.get("success_rate", 0.0),
            resilience_summary.get("timeouts", 0),
            resilience_summary.get("upstream_5xx", 0),
            resilience_summary.get("upstream_429", 0),
            resilience_summary.get("circuit_open_events", 0),
            resilience_summary.get("p50_ms", 0),
            resilience_summary.get("p95_ms", 0)
        ])
    
    # Add compliance_smoke data if available
    if data.get("compliance_smoke") and data["compliance_smoke"].get("summary"):
        compliance_summary = data["compliance_smoke"]["summary"]
        row_data.extend([
            compliance_summary.get("cases_scanned", 0),
            compliance_summary.get("pii_hits", 0),
            compliance_summary.get("rbac_checks", 0),
            compliance_summary.get("rbac_violations", 0),
            compliance_summary.get("pass", False)
        ])
    
    # Add bias_smoke data if available
    if data.get("bias_smoke") and data["bias_smoke"].get("summary"):
        bias_summary = data["bias_smoke"]["summary"]
        row_data.extend([
            bias_summary.get("pairs", 0),
            bias_summary.get("metric", "unknown"),
            bias_summary.get("fails", 0),
            bias_summary.get("fail_ratio", 0.0),
            bias_summary.get("pass", False)
        ])
    
    # Add guardrails data if available
    if data.get("guardrails") and data["guardrails"].get("summary"):
        guardrails_summary = data["guardrails"]["summary"]
        row_data.extend([
            guardrails_summary.get("total", 0),
            guardrails_summary.get("pass", 0),
            guardrails_summary.get("fail", 0),
            guardrails_summary.get("warn", 0),
            guardrails_summary.get("skip", 0)
        ])
    
    # Add Ragas data if available
    rag_summary = summary.get("rag_quality", {})
    if rag_summary.get("ragas"):
        ragas_metrics = rag_summary["ragas"]
        row_data.extend([
            ragas_metrics.get("faithfulness", 0.0),
            ragas_metrics.get("answer_relevancy", 0.0),
            ragas_metrics.get("context_precision", 0.0),
            ragas_metrics.get("context_recall", 0.0),
            rag_summary.get("ragas_thresholds_passed", True)
        ])
    
    # Add Promptfoo data if available
    promptfoo_summary = summary.get("promptfoo", {})
    if promptfoo_summary:
        row_data.extend([
            promptfoo_summary.get("total", 0),
            promptfoo_summary.get("passed", 0),
            promptfoo_summary.get("pass_rate", 0.0)
        ])

    # Add RAG Reliability & Robustness data if available
    rag_reliability_summary = data.get("rag_reliability_robustness", {}).get("prompt_robustness", {}).get("rollups", {})
    if rag_reliability_summary:
        stability_by_mode = rag_reliability_summary.get("stability_avg_by_mode", {})
        row_data.extend([
            rag_reliability_summary.get("capacity_lift_avg", 0.0),
            stability_by_mode.get("simple", 0.0),
            stability_by_mode.get("cot", 0.0),
            stability_by_mode.get("scaffold", 0.0),
            rag_reliability_summary.get("contract_adherence_pct", 0.0)
        ])
    
    # Add Embedding Robustness data if available
    embedding_robustness_data = data.get("rag_reliability_robustness", {}).get("embedding_robustness")
    if embedding_robustness_data:
        row_data.extend([
            embedding_robustness_data.get("avg_recall_at_k", 0.0),
            embedding_robustness_data.get("avg_overlap_at_k", 0.0),
            embedding_robustness_data.get("avg_answer_stability", 0.0),
            embedding_robustness_data.get("low_agreement_count", 0),
            embedding_robustness_data.get("fallback_triggered_count", 0),
            embedding_robustness_data.get("avg_hybrid_gain_delta", 0.0),
            embedding_robustness_data.get("avg_paraphrase_count", 0.0)
        ])
    
    # Add MCP Security data if available
    mcp_summary = summary.get("mcp_security", {})
    if mcp_summary:
        row_data.extend([
            mcp_summary.get("security_tests", 0),
            mcp_summary.get("security_passed", 0),
            mcp_summary.get("p95_latency_ms", 0),
            mcp_summary.get("slo_met", True),
            mcp_summary.get("schema_stable", True),
            mcp_summary.get("out_of_scope_denied", True)
        ])
    
    # Add dataset metadata if available (additive)
    if summary.get("dataset_source"):
        row_data.extend([
            summary.get("dataset_source", "unknown"),
            summary.get("dataset_version", "n/a"),
            summary.get("estimated_tests", 0)
        ])
    
    for col, value in enumerate(row_data, 1):
        ws.cell(row=2, column=col, value=value)
    
    # Add deprecation note if applicable
    if summary.get("_deprecated_note"):
        ws.cell(row=4, column=1, value="Note:")
        ws.cell(row=4, column=2, value=summary["_deprecated_note"])
        ws.cell(row=4, column=1).font = Font(bold=True)
    
    # Freeze panes and format
    ws.freeze_panes = "A2"


def _create_detailed_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Detailed sheet with per-test results."""
    ws = cast(Worksheet, wb.create_sheet("Detailed"))
    
    headers = [
        "suite", "case_id", "provider", "model", "question", "predicted", 
        "expected", "retrieved_count", "recall_at_k", "mrr_at_k", "ndcg_at_k", 
        "faithfulness", "context_recall", "answer_relevancy", "context_precision", 
        "answer_correctness", "answer_similarity", "perturbations_applied", 
        "pass_fail_reason", "latency_ms", "timestamp"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        
        # Set column widths
        if header in ["query_masked", "answer_masked"]:
            ws.column_dimensions[get_column_letter(col)].width = 40  # type: ignore
        elif header in ["metrics_json", "context_ids"]:
            ws.column_dimensions[get_column_letter(col)].width = 30  # type: ignore
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write data rows
    detailed_rows = data.get("detailed", [])
    for row_idx, row_data in enumerate(detailed_rows, 2):
        values = [
            row_data.get("suite", ""),
            row_data.get("case_id", row_data.get("test_id", "")),  # Support both case_id and test_id
            row_data.get("provider", ""),
            row_data.get("model", ""),
            row_data.get("question", row_data.get("query_masked", "")),  # Support both question and query_masked
            row_data.get("predicted", row_data.get("answer_masked", "")),  # Support both predicted and answer_masked
            row_data.get("expected", ""),
            row_data.get("retrieved_count", 0),
            row_data.get("recall_at_k", 0.0),
            row_data.get("mrr_at_k", 0.0),
            row_data.get("ndcg_at_k", 0.0),
            row_data.get("faithfulness", 0.0),
            row_data.get("context_recall", 0.0),
            row_data.get("answer_relevancy", 0.0),
            row_data.get("context_precision", 0.0),
            row_data.get("answer_correctness", 0.0),
            row_data.get("answer_similarity", 0.0),
            row_data.get("perturbations_applied", ""),
            row_data.get("pass_fail_reason", ""),
            row_data.get("latency_ms", 0),
            row_data.get("timestamp", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col in [5, 6]:  # query_masked, answer_masked
                cell.alignment = Alignment(wrap_text=True)  # type: ignore
    
    ws.freeze_panes = "A2"


def _create_api_details_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create API_Details sheet with API call information."""
    ws = cast(Worksheet, wb.create_sheet("API_Details"))
    
    headers = [
        "suite", "test_id", "endpoint", "status_code", "x_source", 
        "x_perf_phase", "x_latency_ms", "request_id", "timestamp"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)  # type: ignore
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write data rows
    api_rows = data.get("api_details", [])
    for row_idx, row_data in enumerate(api_rows, 2):
        values = [
            row_data.get("suite", ""),
            row_data.get("test_id", ""),
            row_data.get("endpoint", ""),
            row_data.get("status_code", ""),
            row_data.get("x_source", ""),
            row_data.get("x_perf_phase", ""),
            row_data.get("x_latency_ms", ""),
            row_data.get("request_id", ""),
            row_data.get("timestamp", "")
        ]
        
        for col, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=value)  # type: ignore
    
    ws.freeze_panes = "A2"


def _create_inputs_expected_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Inputs_And_Expected sheet with test configuration."""
    ws = cast(Worksheet, wb.create_sheet("Inputs_And_Expected"))
    
    headers = [
        "suite", "test_id", "target_mode", "top_k", "options_json", 
        "thresholds_json", "expected_json", "notes", "validation_status",
        "valid_count", "invalid_count", "duplicate_count", "easy_ratio"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        
        # Set column widths for JSON columns
        if "_json" in header:
            ws.column_dimensions[get_column_letter(col)].width = 30  # type: ignore
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write data rows
    inputs_rows = data.get("inputs_expected", [])
    for row_idx, row_data in enumerate(inputs_rows, 2):
        values = [
            row_data.get("suite", ""),
            row_data.get("test_id", ""),
            row_data.get("target_mode", ""),
            row_data.get("top_k", ""),
            json.dumps(row_data.get("options_json", {})) if row_data.get("options_json") else "",
            json.dumps(row_data.get("thresholds_json", {})) if row_data.get("thresholds_json") else "",
            json.dumps(row_data.get("expected_json", {})) if row_data.get("expected_json") else "",
            row_data.get("notes", ""),
            row_data.get("validation_status", ""),
            row_data.get("valid_count", 0),
            row_data.get("invalid_count", 0),
            row_data.get("duplicate_count", 0),
            row_data.get("easy_ratio", 0.0)
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if "_json" in headers[col-1]:
                cell.alignment = Alignment(wrap_text=True)  # type: ignore
    
    ws.freeze_panes = "A2"


def _create_adversarial_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Adversarial_Details sheet for red team results."""
    ws = cast(Worksheet, wb.create_sheet("Adversarial_Details"))
    
    # Required column order as specified
    headers = [
        "run_id", "timestamp", "suite", "provider", "model", "request_id", 
        "attack_id", "attack_text", "response_snippet", "safety_flags", "blocked", "notes"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    # Auto-size columns based on content type
    for col, header in enumerate(headers, 1):
        if header in ["attack_text", "response_snippet"]:
            ws.column_dimensions[get_column_letter(col)].width = 40  # type: ignore
        elif header in ["safety_flags", "notes"]:
            ws.column_dimensions[get_column_letter(col)].width = 25  # type: ignore
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write data rows
    adv_rows = data.get("adversarial_details", [])
    for row_idx, row_data in enumerate(adv_rows, 2):
        values = [
            row_data.get("run_id", ""),
            row_data.get("timestamp", ""),
            row_data.get("suite", ""),
            row_data.get("provider", ""),
            row_data.get("model", ""),
            row_data.get("request_id", ""),
            row_data.get("attack_id", ""),
            row_data.get("attack_text", ""),
            row_data.get("response_snippet", ""),
            json.dumps(row_data.get("safety_flags", [])) if row_data.get("safety_flags") else "",
            row_data.get("blocked", False),
            row_data.get("notes", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            # Apply text wrapping for long content fields
            if headers[col-1] in ["attack_text", "response_snippet", "notes"]:
                cell.alignment = Alignment(wrap_text=True)  # type: ignore
    
    ws.freeze_panes = "A2"


def _create_coverage_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Coverage sheet for test coverage analysis."""
    ws = cast(Worksheet, wb.create_sheet("Coverage"))
    
    # Enhanced headers for RAG coverage analysis
    headers = [
        "module", "stmts", "miss", "branch", "brpart", "cover_percent", "total_lines",
        "total_cases", "filtered_cases", "invalid_cases", "contexts_missing", 
        "perturbed_sampled", "filter_reason"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    # Auto-size columns
    for col, header in enumerate(headers, 1):
        if header == "module":
            ws.column_dimensions[get_column_letter(col)].width = 30  # type: ignore
        else:
            ws.column_dimensions[get_column_letter(col)].width = 12  # type: ignore
    
    # Write data rows from coverage modules
    coverage = data.get("coverage", {})
    modules = coverage.get("modules", [])
    
    for row_idx, module_data in enumerate(modules, 2):
        values = [
            module_data.get("module", ""),
            module_data.get("stmts", 0),
            module_data.get("miss", 0),
            module_data.get("branch", 0),
            module_data.get("brpart", 0),
            module_data.get("cover_percent", 0.0),
            module_data.get("total_lines", 0),
            module_data.get("total_cases", 0),
            module_data.get("filtered_cases", 0),
            module_data.get("invalid_cases", 0),
            module_data.get("contexts_missing", 0),
            module_data.get("perturbed_sampled", 0),
            module_data.get("filter_reason", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
    
    ws.freeze_panes = "A2"


def _create_resilience_details_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Resilience_Details sheet for resilience test results."""
    ws = cast(Worksheet, wb.create_sheet("Resilience_Details"))
    
    # Required column order as specified in requirements
    headers = [
        "run_id", "timestamp", "provider", "model", "request_id", 
        "outcome", "attempts", "latency_ms", "error_class", "mode"
    ]
    
    # Check if we have resilience scenario data to append new columns (additive)
    resilience_data = data.get("resilience", {})
    details = resilience_data.get("details", [])
    has_scenario_data = any(detail.get("scenario_id") for detail in details)
    
    # APPEND new scenario columns at the END (non-breaking)
    if has_scenario_data:
        headers.extend([
            "scenario_id", "failure_mode", "payload_size", "target_timeout_ms", 
            "fail_rate", "circuit_fails", "circuit_reset_s"
        ])
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write resilience detail records
    resilience_details = data.get("resilience", {}).get("details", [])
    
    for row_idx, detail in enumerate(resilience_details, 2):
        values = [
            detail.get("run_id", ""),
            detail.get("timestamp", ""),
            detail.get("provider", ""),
            detail.get("model", ""),
            detail.get("request_id", ""),
            detail.get("outcome", ""),
            detail.get("attempts", 0),
            detail.get("latency_ms", 0),
            detail.get("error_class", ""),
            detail.get("mode", "")
        ]
        
        # APPEND scenario data if present (additive)
        if has_scenario_data:
            values.extend([
                detail.get("scenario_id", ""),
                detail.get("failure_mode", ""),
                detail.get("payload_size", ""),
                detail.get("target_timeout_ms", ""),
                detail.get("fail_rate", ""),
                detail.get("circuit_fails", ""),
                detail.get("circuit_reset_s", "")
            ])
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
    
    ws.freeze_panes = "A2"


def _create_compliance_details_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Compliance_Details sheet for compliance smoke test results."""
    ws = cast(Worksheet, wb.create_sheet("Compliance_Details"))
    
    # Exact column headers as specified
    headers = [
        "run_id", "timestamp", "case_id", "route", "check", "status", "pattern", "notes"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    # Auto-size columns
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write compliance detail records
    compliance_details = data.get("compliance_smoke", {}).get("details", [])
    
    for row_idx, detail in enumerate(compliance_details, 2):
        values = [
            detail.get("run_id", ""),
            detail.get("timestamp", ""),
            detail.get("case_id", ""),
            detail.get("route", ""),
            detail.get("check", ""),
            detail.get("status", ""),
            detail.get("pattern", ""),
            detail.get("notes", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
    
    ws.freeze_panes = "A2"


def _create_bias_details_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Bias_Details sheet for bias smoke test results."""
    ws = cast(Worksheet, wb.create_sheet("Bias_Details"))
    
    # Exact column headers as specified
    headers = [
        "run_id", "timestamp", "case_id", "group_a", "group_b", "metric", "value", "threshold", "question", "answer"
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    # Auto-size columns with special widths for question and answer
    for col in range(1, len(headers) + 1):
        header = headers[col-1]
        if header in ["question", "answer"]:
            ws.column_dimensions[get_column_letter(col)].width = 50
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write bias detail records
    bias_details = data.get("bias_smoke", {}).get("details", [])
    
    for row_idx, detail in enumerate(bias_details, 2):
        values = [
            detail.get("run_id", ""),
            detail.get("timestamp", ""),
            detail.get("case_id", ""),
            detail.get("group_a", ""),
            detail.get("group_b", ""),
            detail.get("metric", ""),
            detail.get("value", ""),
            detail.get("threshold", ""),
            detail.get("question", ""),
            detail.get("answer", "")
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            # Apply text wrapping for long content fields
            if headers[col-1] in ["question", "answer"]:
                cell.alignment = Alignment(wrap_text=True)
    
    ws.freeze_panes = "A2"

def _create_logs_sheet(wb, data):
    """Create Logs sheet with terminal log entries."""
    from openpyxl.worksheet.worksheet import Worksheet
    from openpyxl.styles import Font, Alignment
    from openpyxl.utils import get_column_letter
    from typing import cast
    
    ws = cast(Worksheet, wb.create_sheet("Logs"))
    
    headers = [
        "timestamp", "run_id", "level", "component", "message", 
        "event", "test_id", "provider", "model", "suites"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        
        # Set column widths
        if header == "message":
            ws.column_dimensions[get_column_letter(col)].width = 60  # Wide for messages
        elif header in ["timestamp", "component"]:
            ws.column_dimensions[get_column_letter(col)].width = 20
        elif header == "suites":
            ws.column_dimensions[get_column_letter(col)].width = 25
        else:
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Write log entries
    log_entries = data.get("logs", {}).get("entries", [])
    for row_idx, log_entry in enumerate(log_entries, 2):
        values = [
            log_entry.get("timestamp", ""),
            log_entry.get("run_id", ""),
            log_entry.get("level", ""),
            log_entry.get("component", ""),
            log_entry.get("message", ""),
            log_entry.get("event", ""),
            log_entry.get("test_id", ""),
            log_entry.get("provider", ""),
            log_entry.get("model", ""),
            ",".join(log_entry.get("suites", [])) if isinstance(log_entry.get("suites"), list) else str(log_entry.get("suites", ""))
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            if col == 5:  # message column
                cell.alignment = Alignment(wrap_text=True)
    
    ws.freeze_panes = "A2"


def _create_compare_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Compare sheet with baseline comparison results."""
    ws = cast(Worksheet, wb.create_sheet("Compare"))
    
    compare_data = data.get("compare", {})
    
    # Headers for Compare sheet
    headers = [
        "qid", "question", "primary_model", "baseline_model", "contexts_used_count",
        "answer_sim_primary", "answer_sim_baseline", "Δ_answer_sim",
        "correctness_primary", "correctness_baseline", "Δ_correctness", 
        "faithfulness_primary", "faithfulness_baseline", "Δ_faithfulness",
        "relevancy_primary", "relevancy_baseline", "Δ_relevancy",
        "primary_latency_ms", "baseline_latency_ms", "Δ_latency_ms",
        "baseline_status", "skip_reason", "notes"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        
        # Set column widths based on content
        if header == "question":
            ws.column_dimensions[get_column_letter(col)].width = 40
        elif header in ["primary_model", "baseline_model"]:
            ws.column_dimensions[get_column_letter(col)].width = 20
        elif header in ["skip_reason", "notes"]:
            ws.column_dimensions[get_column_letter(col)].width = 25
        elif "Δ_" in header:
            ws.column_dimensions[get_column_letter(col)].width = 12
        else:
            ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Write comparison cases
    cases = compare_data.get("cases", [])
    for row_idx, case in enumerate(cases, 2):
        # Extract model information
        primary_model = f"{data.get('run', {}).get('provider', 'unknown')}:{data.get('run', {}).get('model', 'unknown')}"
        baseline_model_info = case.get("baseline_model_resolved", {})
        baseline_model = f"{baseline_model_info.get('preset', 'unknown')}:{baseline_model_info.get('model', 'unknown')}"
        
        # Extract metrics
        primary_metrics = case.get("primary_metrics", {})
        baseline_metrics = case.get("baseline_metrics", {})
        delta_metrics = case.get("delta_metrics", {})
        
        # Helper function to get metric value or N/A
        def get_metric(metrics_dict, key, default="N/A"):
            return metrics_dict.get(key, default) if metrics_dict else default
        
        values = [
            case.get("qid", ""),
            case.get("question", "")[:100] + "..." if len(case.get("question", "")) > 100 else case.get("question", ""),
            primary_model,
            baseline_model,
            case.get("primary_contexts_used_count", 0),
            get_metric(primary_metrics, "answer_similarity"),
            get_metric(baseline_metrics, "answer_similarity"),
            get_metric(delta_metrics, "answer_similarity"),
            get_metric(primary_metrics, "answer_correctness"),
            get_metric(baseline_metrics, "answer_correctness"),
            get_metric(delta_metrics, "answer_correctness"),
            get_metric(primary_metrics, "faithfulness"),
            get_metric(baseline_metrics, "faithfulness"),
            get_metric(delta_metrics, "faithfulness"),
            get_metric(primary_metrics, "answer_relevancy"),
            get_metric(baseline_metrics, "answer_relevancy"),
            get_metric(delta_metrics, "answer_relevancy"),
            case.get("primary_latency_ms", "N/A"),
            case.get("baseline_latency_ms", "N/A"),
            (case.get("baseline_latency_ms", 0) - case.get("primary_latency_ms", 0)) if case.get("baseline_latency_ms") and case.get("primary_latency_ms") else "N/A",
            case.get("baseline_status", ""),
            case.get("skip_reason", ""),
            case.get("error", "") or ""
        ]
        
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            
            # Apply formatting for delta columns (highlight positive/negative)
            if isinstance(value, (int, float)) and "Δ_" in headers[col-1]:
                if value > 0:
                    cell.font = Font(color="00AA00")  # Green for positive deltas
                elif value < 0:
                    cell.font = Font(color="AA0000")  # Red for negative deltas
    
    # Add summary section at the top (insert rows)
    ws.insert_rows(1, 6)
    
    # Summary header
    ws.cell(row=1, column=1, value="Compare (contexts-only)").font = Font(bold=True, size=14)
    
    # Summary statistics
    summary = compare_data.get("summary", {})
    aggregates = compare_data.get("aggregates", {})
    
    summary_info = [
        f"Compared cases: {summary.get('compared_cases', 0)}",
        f"Skipped (no contexts): {summary.get('skipped_no_contexts', 0)}",
        f"Skipped (missing creds): {summary.get('skipped_missing_creds', 0)}",
        f"Skipped (no candidate): {summary.get('skipped_no_candidate', 0)}",
        f"Skipped (errors): {summary.get('skipped_errors', 0)}"
    ]
    
    for i, info in enumerate(summary_info, 2):
        ws.cell(row=i, column=1, value=info)
    
    # Add aggregate metrics if available
    if aggregates and summary.get('compared_cases', 0) > 0:
        ws.cell(row=2, column=3, value="Average Metrics:")
        ws.cell(row=2, column=3).font = Font(bold=True)
        
        metric_row = 3
        for key, value in aggregates.items():
            if isinstance(value, (int, float)):
                ws.cell(row=metric_row, column=3, value=f"{key}: {value:.3f}")
                metric_row += 1
    elif summary.get('compared_cases', 0) == 0:
        ws.cell(row=2, column=3, value="No contexts observed; comparison skipped.")
    
    # Adjust header row (now row 7 after inserts)
    header_row = 7
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
    
    ws.freeze_panes = f"A{header_row + 1}"


def _create_guardrails_details_sheet(wb: Workbook, data: Dict[str, Any]) -> None:
    """Create Guardrails_Details sheet with guardrails test results."""
    ws = cast(Worksheet, wb.create_sheet("Guardrails_Details"))
    
    headers = [
        "subtest", "suite", "test_id", "status", "reason", "thresholds", 
        "links", "category", "description", "enabled", "latency_ms", "timestamp"
    ]
    
    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)
        
        # Set column widths
        if header in ["reason", "description"]:
            ws.column_dimensions[get_column_letter(col)].width = 40
        elif header in ["links", "thresholds"]:
            ws.column_dimensions[get_column_letter(col)].width = 30
        else:
            ws.column_dimensions[get_column_letter(col)].width = 16
    
    # Write data rows
    guardrails_details = data.get("guardrails", {}).get("details", [])
    for row_idx, detail in enumerate(guardrails_details, 2):
        values = [
            detail.get("subtest", ""),
            detail.get("suite", ""),
            detail.get("test_id", ""),
            detail.get("status", ""),
            detail.get("reason", ""),
            detail.get("thresholds", ""),
            detail.get("links", ""),
            detail.get("category", ""),
            detail.get("description", ""),
            detail.get("enabled", ""),
            detail.get("latency_ms", ""),
            detail.get("timestamp", "")
        ]
        
        for col, value in enumerate(values, 1):
            ws.cell(row=row_idx, column=col, value=value)
    
    # Freeze panes
    ws.freeze_panes = "A2"

