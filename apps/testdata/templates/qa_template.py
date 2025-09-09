"""Excel template generator for QA sets and passages."""

import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def create_qa_template() -> bytes:
    """Create Excel template for QA set with proper formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "QA Template"
    
    # Headers - Match JSONL format exactly
    headers = ["QID", "Question", "Expected Answer", "Contexts", "Meta Category", "Meta Difficulty"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data - Match JSONL format exactly
    sample_data = [
        [
            "q1",  # QID
            "What is the capital of France?",  # Question
            "Paris",  # Expected Answer
            "p1",  # Contexts (passage ID)
            "geography",  # Meta Category
            "easy"  # Meta Difficulty
        ],
        [
            "q2",
            "How do you calculate compound interest?",
            "A = P(1 + r/n)^(nt)",
            "p2",
            "finance",
            "medium"
        ],
        [
            "q3",
            "What is machine learning?",
            "A subset of AI that enables computers to learn from data",
            "p3",
            "technology",
            "easy"
        ]
    ]
    
    # Add sample data
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        ws.column_dimensions[column_letter].width = 25
    
    # Add instructions
    ws.cell(row=6, column=1, value="Instructions:")
    ws.cell(row=7, column=1, value="1. Replace sample data with your own questions and answers")
    ws.cell(row=8, column=1, value="2. Question: The question you want to test")
    ws.cell(row=9, column=1, value="3. Context: Relevant background information (optional)")
    ws.cell(row=10, column=1, value="4. Expected Answer: The correct answer you expect")
    ws.cell(row=11, column=1, value="5. Metadata: Additional categorization (optional)")
    
    # Format instructions
    for row in range(6, 12):
        cell = ws.cell(row=row, column=1)
        if row == 6:
            cell.font = Font(bold=True)
        else:
            cell.font = Font(italic=True)
    
    # Save to bytes
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def create_passages_template() -> bytes:
    """Create Excel template for passages with proper formatting."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Passages Template"
    
    # Headers - Match JSONL format exactly
    headers = ["ID", "Text", "Meta Source", "Meta Category"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Sample data - Match JSONL format exactly
    sample_data = [
        [
            "p1",  # ID
            "France is a country in Western Europe. Paris is the capital and largest city of France, with a population of over 2 million people.",  # Text
            "geography_textbook",  # Meta Source
            "geography"  # Meta Category
        ],
        [
            "p2", 
            "Compound interest is the interest calculated on the initial principal and also on the accumulated interest of previous periods. The formula is A = P(1 + r/n)^(nt).",
            "finance_guide",
            "finance"
        ],
        [
            "p3",
            "Machine learning is a subset of artificial intelligence that enables computers to learn and improve from experience without being explicitly programmed.",
            "tech_encyclopedia",
            "technology"
        ]
    ]
    
    # Add sample data
    for row, data in enumerate(sample_data, 2):
        for col, value in enumerate(data, 1):
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    ws.column_dimensions['A'].width = 10  # ID column
    ws.column_dimensions['B'].width = 50  # Text column
    ws.column_dimensions['C'].width = 15  # Metadata column
    
    # Add instructions
    ws.cell(row=6, column=1, value="Instructions:")
    ws.cell(row=7, column=1, value="1. Replace sample data with your own text passages")
    ws.cell(row=8, column=1, value="2. ID: Unique identifier for each passage (e.g., p1, p2, p3)")
    ws.cell(row=9, column=1, value="3. Text: The actual text content of the passage")
    ws.cell(row=10, column=1, value="4. Metadata: Additional categorization or tags (optional)")
    
    # Format instructions
    for row in range(6, 11):
        cell = ws.cell(row=row, column=1)
        if row == 6:
            cell.font = Font(bold=True)
        else:
            cell.font = Font(italic=True)
    
    # Save to bytes
    from io import BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


def create_qa_jsonl_template() -> str:
    """Create JSONL template for QA set."""
    sample_data = [
        {
            "qid": "q1",
            "question": "What is the capital of France?",
            "expected_answer": "Paris",
            "contexts": ["p1"],
            "meta": {"category": "geography", "difficulty": "easy"}
        },
        {
            "qid": "q2", 
            "question": "How do you calculate compound interest?",
            "expected_answer": "A = P(1 + r/n)^(nt)",
            "contexts": ["p2"],
            "meta": {"category": "finance", "difficulty": "medium"}
        },
        {
            "qid": "q3",
            "question": "What is machine learning?", 
            "expected_answer": "A subset of AI that enables computers to learn from data",
            "contexts": ["p3"],
            "meta": {"category": "technology", "difficulty": "easy"}
        }
    ]
    
    lines = []
    lines.append("// QA Set JSONL Template")
    lines.append("// Each line should be a valid JSON object with the following fields:")
    lines.append("// - qid: Unique question identifier")
    lines.append("// - question: The question text")
    lines.append("// - expected_answer: The correct answer")
    lines.append("// - contexts: Array of passage IDs (optional)")
    lines.append("// - meta: Additional metadata (optional)")
    lines.append("")
    
    for item in sample_data:
        lines.append(json.dumps(item, ensure_ascii=False))
    
    return "\n".join(lines)


def create_passages_jsonl_template() -> str:
    """Create JSONL template for passages."""
    sample_data = [
        {
            "id": "p1",
            "text": "France is a country in Western Europe. Paris is the capital and largest city of France, with a population of over 2 million people.",
            "meta": {"source": "geography_textbook", "category": "geography"}
        },
        {
            "id": "p2",
            "text": "Compound interest is the interest calculated on the initial principal and also on the accumulated interest of previous periods. The formula is A = P(1 + r/n)^(nt).",
            "meta": {"source": "finance_guide", "category": "finance"}
        },
        {
            "id": "p3", 
            "text": "Machine learning is a subset of artificial intelligence that enables computers to learn and improve from experience without being explicitly programmed.",
            "meta": {"source": "tech_encyclopedia", "category": "technology"}
        }
    ]
    
    lines = []
    lines.append("// Passages JSONL Template")
    lines.append("// Each line should be a valid JSON object with the following fields:")
    lines.append("// - id: Unique passage identifier")
    lines.append("// - text: The passage content")
    lines.append("// - meta: Additional metadata (optional)")
    lines.append("")
    
    for item in sample_data:
        lines.append(json.dumps(item, ensure_ascii=False))
    
    return "\n".join(lines)
