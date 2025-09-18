"""Excel to JSONL conversion utilities for RAG test data."""

import json
import logging
from typing import List, Dict, Any, Tuple, Optional
from pathlib import Path
import tempfile
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)


def to_jsonl_qaset(excel_file_path: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Convert Excel QA template to JSONL format.
    
    Expected columns: Question, Context, Expected Answer, Metadata (optional)
    
    Args:
        excel_file_path: Path to Excel file
        
    Returns:
        Tuple of (jsonl_content_string, parsed_records_list)
    """
    try:
        workbook = load_workbook(excel_file_path, read_only=True)
        worksheet = workbook.active
        
        # Find header row (usually row 1)
        headers = []
        if worksheet is None:
            raise ValueError("Excel file has no active worksheet")
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append("")
        
        # Map expected column names to indices
        column_map = {}
        for i, header in enumerate(headers):
            if header in ["qid", "id"]:
                column_map["qid"] = i
            elif "question" in header:
                column_map["question"] = i
            elif "context" in header:
                column_map["contexts"] = i
            elif "expected" in header and "answer" in header:
                column_map["expected_answer"] = i
            elif "answer" in header and "expected" not in header:
                column_map["expected_answer"] = i
            elif "meta" in header and "category" in header:
                column_map["meta_category"] = i
            elif "meta" in header and "difficulty" in header:
                column_map["meta_difficulty"] = i
            elif "metadata" in header or "meta" in header:
                column_map["metadata"] = i
            # New robustness fields
            elif "required" in header:
                column_map["required"] = i
            elif "task" in header and "type" in header:
                column_map["task_type"] = i
            elif "robustness" in header and "paraphrase" in header:
                column_map["robustness_paraphrases"] = i
            elif "robustness" in header and "synonym" in header:
                column_map["robustness_synonyms"] = i
            elif "prompt" in header and "robustness" in header and "enabled" in header:
                column_map["prompt_robustness_enabled"] = i
        
        # Validate required columns
        required_cols = ["question", "expected_answer"]
        missing_cols = [col for col in required_cols if col not in column_map]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        
        records = []
        jsonl_lines = []
        
        # Process data rows (skip header)
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue
            
            # Extract values
            question = row[column_map["question"]] if column_map["question"] < len(row) else None
            expected_answer = row[column_map["expected_answer"]] if column_map["expected_answer"] < len(row) else None
            
            # Skip rows with missing required data
            if not question or not expected_answer:
                continue
            
            # Extract QID (use provided or generate)
            qid = None
            if "qid" in column_map and column_map["qid"] < len(row):
                qid = row[column_map["qid"]]
            if not qid:
                qid = f"q{row_num-1}"
            
            # Build record
            record: Dict[str, Any] = {
                "qid": str(qid).strip(),
                "question": str(question).strip(),
                "expected_answer": str(expected_answer).strip()
            }
            
            # Add contexts if present
            if "contexts" in column_map and column_map["contexts"] < len(row):
                contexts = row[column_map["contexts"]]
                if contexts:
                    # Handle both single ID and comma-separated IDs
                    context_ids = [c.strip() for c in str(contexts).split(",")]
                    record["contexts"] = context_ids
            
            # Add robustness fields if present
            if "required" in column_map and column_map["required"] < len(row):
                required_val = row[column_map["required"]]
                if required_val and str(required_val).strip().upper() in ["TRUE", "1", "YES"]:
                    record["required"] = True
            
            if "task_type" in column_map and column_map["task_type"] < len(row):
                task_type = row[column_map["task_type"]]
                if task_type:
                    record["task_type"] = str(task_type).strip()
            
            # Build robustness config if any robustness fields are present
            robustness_config = {}
            if "robustness_paraphrases" in column_map and column_map["robustness_paraphrases"] < len(row):
                paraphrases = row[column_map["robustness_paraphrases"]]
                if paraphrases:
                    # Split by semicolon and clean
                    paraphrase_list = [p.strip() for p in str(paraphrases).split(";") if p.strip()]
                    if paraphrase_list:
                        robustness_config["paraphrases"] = paraphrase_list
            
            if "robustness_synonyms" in column_map and column_map["robustness_synonyms"] < len(row):
                synonyms = row[column_map["robustness_synonyms"]]
                if synonyms:
                    # Split by semicolon and clean
                    synonym_list = [s.strip() for s in str(synonyms).split(";") if s.strip()]
                    if synonym_list:
                        robustness_config["synonyms"] = synonym_list
            
            if robustness_config:
                record["robustness"] = robustness_config
            
            # Build prompt robustness config if enabled
            if "prompt_robustness_enabled" in column_map and column_map["prompt_robustness_enabled"] < len(row):
                pr_enabled = row[column_map["prompt_robustness_enabled"]]
                if pr_enabled and str(pr_enabled).strip().upper() in ["TRUE", "1", "YES"]:
                    record["prompt_robustness"] = {
                        "enabled": True,
                        "modes": ["simple", "cot", "scaffold"],
                        "paraphrase_runs": 2
                    }
            
            # Build metadata object
            meta = {}
            
            # Add category if present
            if "meta_category" in column_map and column_map["meta_category"] < len(row):
                category = row[column_map["meta_category"]]
                if category:
                    meta["category"] = str(category).strip()
            
            # Add difficulty if present
            if "meta_difficulty" in column_map and column_map["meta_difficulty"] < len(row):
                difficulty = row[column_map["meta_difficulty"]]
                if difficulty:
                    meta["difficulty"] = str(difficulty).strip()
            
            # Add legacy metadata if present
            if "metadata" in column_map and column_map["metadata"] < len(row):
                metadata = row[column_map["metadata"]]
                if metadata:
                    try:
                        # Try to parse as JSON, otherwise use as string
                        legacy_meta = json.loads(str(metadata))
                        meta.update(legacy_meta)
                    except:
                        meta["category"] = str(metadata)
            
            # Add meta to record if not empty
            if meta:
                record["meta"] = meta
            
            records.append(record)
            jsonl_lines.append(json.dumps(record))
        
        jsonl_content = "\n".join(jsonl_lines)
        logger.info(f"Converted Excel to {len(records)} QA records")
        
        return jsonl_content, records
        
    except Exception as e:
        logger.error(f"Error converting Excel QA set: {e}")
        raise ValueError(f"Failed to convert Excel QA set: {str(e)}")


def to_jsonl_passages(excel_file_path: str) -> Tuple[str, List[Dict[str, Any]]]:
    """
    Convert Excel passages template to JSONL format.
    
    Expected columns: ID (optional), Text, Metadata (optional)
    
    Args:
        excel_file_path: Path to Excel file
        
    Returns:
        Tuple of (jsonl_content_string, parsed_records_list)
    """
    try:
        workbook = load_workbook(excel_file_path, read_only=True)
        worksheet = workbook.active
        
        # Find header row
        headers = []
        if worksheet is None:
            raise ValueError("Excel file has no active worksheet")
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
            else:
                headers.append("")
        
        # Map column names to indices
        column_map = {}
        for i, header in enumerate(headers):
            if header in ["id", "passage_id", "doc_id"]:
                column_map["id"] = i
            elif "text" in header or "content" in header or "passage" in header:
                column_map["text"] = i
            elif "meta" in header and "source" in header:
                column_map["meta_source"] = i
            elif "meta" in header and "category" in header:
                column_map["meta_category"] = i
            elif "metadata" in header or "meta" in header:
                column_map["metadata"] = i
        
        # Validate required columns
        if "text" not in column_map:
            raise ValueError("Missing required 'text' column")
        
        records = []
        jsonl_lines = []
        
        # Process data rows
        for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):  # Skip empty rows
                continue
            
            # Extract text
            text = row[column_map["text"]] if column_map["text"] < len(row) else None
            if not text:
                continue
            
            # Build record
            record: Dict[str, Any] = {
                "text": str(text).strip()
            }
            
            # Add ID if present, otherwise generate
            if "id" in column_map and column_map["id"] < len(row):
                passage_id = row[column_map["id"]]
                if passage_id:
                    record["id"] = str(passage_id).strip()
                else:
                    record["id"] = f"p{row_num-1}"
            else:
                record["id"] = f"p{row_num-1}"
            
            # Build metadata object
            meta = {}
            
            # Add source if present
            if "meta_source" in column_map and column_map["meta_source"] < len(row):
                source = row[column_map["meta_source"]]
                if source:
                    meta["source"] = str(source).strip()
            
            # Add category if present
            if "meta_category" in column_map and column_map["meta_category"] < len(row):
                category = row[column_map["meta_category"]]
                if category:
                    meta["category"] = str(category).strip()
            
            # Add legacy metadata if present
            if "metadata" in column_map and column_map["metadata"] < len(row):
                metadata = row[column_map["metadata"]]
                if metadata:
                    try:
                        legacy_meta = json.loads(str(metadata))
                        meta.update(legacy_meta)
                    except:
                        meta["source"] = str(metadata)
            
            # Add meta to record if not empty
            if meta:
                record["meta"] = meta
            
            records.append(record)
            jsonl_lines.append(json.dumps(record))
        
        jsonl_content = "\n".join(jsonl_lines)
        logger.info(f"Converted Excel to {len(records)} passage records")
        
        return jsonl_content, records
        
    except Exception as e:
        logger.error(f"Error converting Excel passages: {e}")
        raise ValueError(f"Failed to convert Excel passages: {str(e)}")


def detect_excel_type(excel_file_path: str) -> str:
    """
    Detect whether Excel file is QA set or passages based on column headers.
    
    Args:
        excel_file_path: Path to Excel file
        
    Returns:
        "qaset" or "passages"
    """
    try:
        workbook = load_workbook(excel_file_path, read_only=True)
        worksheet = workbook.active
        
        # Get first row headers
        headers = []
        if worksheet is None:
            raise ValueError("Excel file has no active worksheet")
        for cell in worksheet[1]:
            if cell.value:
                headers.append(str(cell.value).strip().lower())
        
        header_text = " ".join(headers)
        
        # Check for QA set indicators
        qa_indicators = ["question", "expected", "answer"]
        passage_indicators = ["text", "content", "passage"]
        
        qa_score = sum(1 for indicator in qa_indicators if indicator in header_text)
        passage_score = sum(1 for indicator in passage_indicators if indicator in header_text)
        
        if qa_score > passage_score:
            return "qaset"
        else:
            return "passages"
            
    except Exception as e:
        logger.warning(f"Could not detect Excel type: {e}, defaulting to qaset")
        return "qaset"


def convert_excel_file(file_path: str, target_type: Optional[str] = None) -> Tuple[str, str, List[Dict[str, Any]]]:
    """
    Convert Excel file to JSONL based on detected or specified type.
    
    Args:
        file_path: Path to Excel file
        target_type: Force conversion type ("qaset" or "passages"), or None for auto-detect
        
    Returns:
        Tuple of (detected_type, jsonl_content, records_list)
    """
    if target_type:
        conversion_type = target_type
    else:
        conversion_type = detect_excel_type(file_path)
    
    if conversion_type == "qaset":
        jsonl_content, records = to_jsonl_qaset(file_path)
    else:
        jsonl_content, records = to_jsonl_passages(file_path)
    
    return conversion_type, jsonl_content, records
