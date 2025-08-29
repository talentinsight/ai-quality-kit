#!/usr/bin/env python3
"""End-to-end smoke tests for compliance_smoke and bias_smoke suites."""

import asyncio
import json
import sys
import tempfile
from pathlib import Path
import httpx
import pytest

# Add parent directory to Python path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from apps.orchestrator.run_tests import TestRunner, OrchestratorRequest


@pytest.mark.asyncio
async def test_compliance_smoke_e2e():
    """Test compliance_smoke suite end-to-end."""
    print("🔍 Testing compliance_smoke suite...")
    
    # Create temporary PII patterns file
    patterns = {
        "ssn": {
            "pattern": "\\b\\d{3}-\\d{2}-\\d{4}\\b",
            "description": "Social Security Number"
        },
        "email": {
            "pattern": "\\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\\.[A-Z|a-z]{2,}\\b",
            "description": "Email address"
        }
    }
    
    with tempfile.NamedTemporaryFile(mode='w', suffix='.json', delete=False) as f:
        json.dump(patterns, f)
        patterns_file = f.name
    
    try:
        request = OrchestratorRequest(
            target_mode="api",
            suites=["compliance_smoke"],
            options={
                "provider": "mock",
                "model": "mock-model",
                "compliance_smoke": {
                    "pii_patterns_file": patterns_file,
                    "scan_fields": ["answer"],
                    "rbac_matrix": {
                        "user": ["/ask", "/orchestrator/*"],
                        "admin": ["*"]
                    }
                }
            }
        )
        
        runner = TestRunner(request)
        
        # Load tests
        suite_data = runner.load_suites()
        print(f"  ✓ Loaded {len(suite_data.get('compliance_smoke', []))} compliance tests")
        
        # Check test structure
        compliance_tests = suite_data.get('compliance_smoke', [])
        pii_tests = [t for t in compliance_tests if t.get('category') == 'pii_scan']
        rbac_tests = [t for t in compliance_tests if t.get('category') == 'rbac_probe']
        
        print(f"  ✓ PII tests: {len(pii_tests)}")
        print(f"  ✓ RBAC tests: {len(rbac_tests)}")
        
        # Verify test structure
        assert len(pii_tests) > 0, "Should have PII tests"
        assert len(rbac_tests) > 0, "Should have RBAC tests"
        
        for test in compliance_tests:
            assert 'test_id' in test
            assert 'compliance_config' in test
            assert test['test_type'] == 'compliance_smoke'
        
        print("  ✅ Compliance smoke suite structure verified")
        return True
        
    finally:
        # Clean up
        Path(patterns_file).unlink(missing_ok=True)


@pytest.mark.asyncio
async def test_bias_smoke_e2e():
    """Test bias_smoke suite end-to-end."""
    print("⚖️  Testing bias_smoke suite...")
    
    request = OrchestratorRequest(
        target_mode="api",
        suites=["bias_smoke"],
        options={
            "provider": "mock",
            "model": "mock-model",
            "bias_smoke": {
                "groups": [["female", "male"], ["young", "elderly"]],
                "max_pairs": 3,
                "parity_metric": "refusal_rate",
                "parity_threshold": 0.25
            }
        }
    )
    
    runner = TestRunner(request)
    
    # Load tests
    suite_data = runner.load_suites()
    bias_tests = suite_data.get('bias_smoke', [])
    
    print(f"  ✓ Loaded {len(bias_tests)} bias tests")
    
    # Verify test structure
    assert len(bias_tests) <= 3, "Should respect max_pairs limit"
    assert len(bias_tests) > 0, "Should have bias tests"
    
    for test in bias_tests:
        assert 'test_id' in test
        assert 'query_a' in test
        assert 'query_b' in test
        assert 'bias_config' in test
        assert test['test_type'] == 'bias_smoke'
        assert test['category'] == 'demographic_parity'
        
        config = test['bias_config']
        assert config['group_a'] in ['female', 'young']
        assert config['group_b'] in ['male', 'elderly']
        assert config['parity_metric'] == 'refusal_rate'
        assert config['parity_threshold'] == 0.25
    
    print("  ✅ Bias smoke suite structure verified")
    return True


@pytest.mark.asyncio
async def test_sharding_e2e():
    """Test sharding functionality."""
    print("🗂️  Testing sharding functionality...")
    
    request = OrchestratorRequest(
        target_mode="api",
        suites=["compliance_smoke", "bias_smoke"],
        shards=2,
        shard_id=1,
        options={
            "provider": "mock",
            "model": "mock-model"
        }
    )
    
    runner = TestRunner(request)
    
    # Load tests with sharding
    suite_data = runner.load_suites()
    
    print(f"  ✓ Compliance tests in shard: {len(suite_data.get('compliance_smoke', []))}")
    print(f"  ✓ Bias tests in shard: {len(suite_data.get('bias_smoke', []))}")
    
    # Verify sharding applied
    total_tests = sum(len(tests) for tests in suite_data.values())
    assert total_tests > 0, "Should have some tests in shard"
    
    # Test deterministic sharding
    runner2 = TestRunner(request)
    suite_data2 = runner2.load_suites()
    
    assert suite_data == suite_data2, "Sharding should be deterministic"
    
    print("  ✅ Sharding functionality verified")
    return True


@pytest.mark.asyncio
async def test_json_reporting_e2e():
    """Test JSON reporting for smoke suites."""
    print("📄 Testing JSON reporting...")
    
    from apps.reporters.json_reporter import build_json
    
    # Mock data
    run_meta = {
        "run_id": "test_run_123",
        "started_at": "2024-01-01T00:00:00Z",
        "finished_at": "2024-01-01T00:05:00Z",
        "provider": "mock",
        "model": "mock-model"
    }
    
    summary = {
        "overall": {"total_tests": 5, "passed": 3, "failed": 2},
        "compliance_smoke": {
            "cases_scanned": 4,
            "pii_hits": 1,
            "rbac_checks": 2,
            "rbac_violations": 0,
            "pass": False
        },
        "bias_smoke": {
            "pairs": 3,
            "metric": "refusal_rate",
            "fails": 1,
            "fail_ratio": 0.33,
            "pass": False
        }
    }
    
    compliance_details = [
        {
            "run_id": "test_run_123",
            "timestamp": "2024-01-01T00:01:00Z",
            "case_id": "compliance_pii_1",
            "route": "N/A",
            "check": "pii",
            "status": "fail",
            "pattern": "ssn",
            "notes": "Found 1 PII matches"
        }
    ]
    
    bias_details = [
        {
            "run_id": "test_run_123",
            "timestamp": "2024-01-01T00:02:00Z",
            "case_id": "bias_pair_1",
            "group_a": "female",
            "group_b": "male",
            "metric": "refusal_rate",
            "value": 0.5,
            "threshold": 0.25,
            "notes": "Test comparison"
        }
    ]
    
    # Build JSON report
    report = build_json(
        run_meta=run_meta,
        summary=summary,
        detailed_rows=[],
        api_rows=[],
        inputs_rows=[],
        compliance_smoke_details=compliance_details,
        bias_smoke_details=bias_details
    )
    
    # Verify structure
    assert "compliance_smoke" in report
    assert "bias_smoke" in report
    assert "summary" in report["compliance_smoke"]
    assert "details" in report["compliance_smoke"]
    assert "summary" in report["bias_smoke"]
    assert "details" in report["bias_smoke"]
    
    # Verify content
    assert report["compliance_smoke"]["summary"]["pass"] == False
    assert report["bias_smoke"]["summary"]["pass"] == False
    assert len(report["compliance_smoke"]["details"]) == 1
    assert len(report["bias_smoke"]["details"]) == 1
    
    print("  ✅ JSON reporting structure verified")
    return True


@pytest.mark.asyncio
async def test_excel_reporting_e2e():
    """Test Excel reporting for smoke suites."""
    print("📊 Testing Excel reporting...")
    
    from apps.reporters.excel_reporter import _create_compliance_details_sheet, _create_bias_details_sheet
    from openpyxl import Workbook
    from openpyxl.worksheet.worksheet import Worksheet
    from typing import cast
    import tempfile
    
    # Test compliance sheet
    wb = Workbook()
    data = {
        "compliance_smoke": {
            "details": [
                {
                    "run_id": "test_run_123",
                    "timestamp": "2024-01-01T00:01:00Z",
                    "case_id": "compliance_pii_1",
                    "route": "N/A",
                    "check": "pii",
                    "status": "fail",
                    "pattern": "ssn",
                    "notes": "Found PII"
                }
            ]
        }
    }
    
    _create_compliance_details_sheet(wb, data)
    
    # Verify sheet exists and has correct structure
    assert "Compliance_Details" in wb.sheetnames
    ws = cast(Worksheet, wb["Compliance_Details"])
    
    # Check headers
    expected_headers = ["run_id", "timestamp", "case_id", "route", "check", "status", "pattern", "notes"]
    for col, header in enumerate(expected_headers, 1):
        assert ws.cell(row=1, column=col).value == header
    
    # Check data
    assert ws.cell(row=2, column=1).value == "test_run_123"
    assert ws.cell(row=2, column=5).value == "pii"
    assert ws.cell(row=2, column=6).value == "fail"
    
    print("  ✓ Compliance_Details sheet verified")
    
    # Test bias sheet
    wb2 = Workbook()
    data2 = {
        "bias_smoke": {
            "details": [
                {
                    "run_id": "test_run_123",
                    "timestamp": "2024-01-01T00:02:00Z",
                    "case_id": "bias_pair_1",
                    "group_a": "female",
                    "group_b": "male",
                    "metric": "refusal_rate",
                    "value": 0.5,
                    "threshold": 0.25,
                    "notes": "Test note"
                }
            ]
        }
    }
    
    _create_bias_details_sheet(wb2, data2)
    
    # Verify sheet exists and has correct structure
    assert "Bias_Details" in wb2.sheetnames
    ws2 = cast(Worksheet, wb2["Bias_Details"])
    
    # Check headers
    expected_headers2 = ["run_id", "timestamp", "case_id", "group_a", "group_b", "metric", "value", "threshold", "notes"]
    for col, header in enumerate(expected_headers2, 1):
        assert ws2.cell(row=1, column=col).value == header
    
    # Check data
    assert ws2.cell(row=2, column=1).value == "test_run_123"
    assert ws2.cell(row=2, column=4).value == "female"
    assert ws2.cell(row=2, column=5).value == "male"
    assert ws2.cell(row=2, column=6).value == "refusal_rate"
    
    print("  ✓ Bias_Details sheet verified")
    print("  ✅ Excel reporting structure verified")
    return True


async def main():
    """Run all smoke suite tests."""
    print("🚀 Running smoke suites end-to-end tests...\n")
    
    tests = [
        test_compliance_smoke_e2e,
        test_bias_smoke_e2e,
        test_sharding_e2e,
        test_json_reporting_e2e,
        test_excel_reporting_e2e
    ]
    
    passed = 0
    failed = 0
    
    for test_func in tests:
        try:
            await test_func()
            passed += 1
        except Exception as e:
            print(f"  ❌ {test_func.__name__} failed: {e}")
            failed += 1
        print()
    
    print(f"📊 Results: {passed} passed, {failed} failed")
    
    if failed > 0:
        print("❌ Some tests failed!")
        return 1
    else:
        print("✅ All smoke suite tests passed!")
        return 0


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    sys.exit(exit_code)
