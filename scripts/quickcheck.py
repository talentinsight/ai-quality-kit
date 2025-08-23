#!/usr/bin/env python3
"""
Quick Validation Pack: start (optional) the API, run smoke checks,
exercise /ask, orchestrator run, download JSON/XLSX reports, validate sheets.
Zero-retention by default; uses provider=mock, no secrets printed.
"""

import os, sys, time, json, subprocess, signal, tempfile, pathlib
from typing import Optional, Tuple, Dict, Any, List

import httpx

REQUIRED_SHEETS = [
    "Summary", "Detailed", "API_Details", "Inputs_And_Expected"
    # "Adversarial_Details", "Coverage" appear when red_team is selected
]

PORT = int(os.getenv("QUICKCHECK_PORT", "8000"))
BASE = os.getenv("QUICKCHECK_BASE", f"http://localhost:{PORT}")
TOKEN = os.getenv("QUICKCHECK_TOKEN", "SECRET_USER")
START_SERVER = os.getenv("QUICKCHECK_START_SERVER", "true").lower() in {"1","true","yes"}
SUITES = os.getenv("QUICKCHECK_SUITES", "rag_quality,performance").split(",")

def _print_ok(msg: str) -> None:
    print(f"[OK] {msg}")

def _print_info(msg: str) -> None:
    print(f"[..] {msg}")

def _print_fail(msg: str) -> None:
    print(f"[FAIL] {msg}")

def wait_ready(timeout_s: int = 60) -> None:
    _print_info("Waiting for /readyz ...")
    deadline = time.time() + timeout_s
    with httpx.Client(timeout=5.0) as cx:
        while time.time() < deadline:
            try:
                r = cx.get(f"{BASE}/readyz")
                if r.status_code == 200:
                    _print_ok("Service ready.")
                    return
            except Exception:
                pass
            time.sleep(0.5)
    raise RuntimeError("Service did not become ready in time.")

def start_server() -> Optional[subprocess.Popen]:
    if not START_SERVER:
        _print_info("Skipping server start (QUICKCHECK_START_SERVER=false).")
        return None
    _print_info("Starting uvicorn apps.rag_service.main:app ...")
    env = os.environ.copy()
    # zero-retention & auth defaults for safety
    env.setdefault("AUTH_ENABLED", "true")
    env.setdefault("AUTH_TOKENS", "admin:SECRET_ADMIN,user:SECRET_USER")
    env.setdefault("PERSIST_DB", "false")
    env.setdefault("ANONYMIZE_REPORTS", "true")
    env.setdefault("MCP_ENABLED", "true")
    env.setdefault("A2A_ENABLED", "true")
    # prefer unbuffered logs for readiness
    cmd = [
        sys.executable, "-m", "uvicorn",
        "apps.rag_service.main:app", "--port", str(PORT), "--reload"
    ]
    proc = subprocess.Popen(cmd, env=env, stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
    # small grace period
    time.sleep(0.5)
    return proc

def stop_server(proc: Optional[subprocess.Popen]) -> None:
    if proc is None:
        return
    _print_info("Stopping server ...")
    try:
        proc.terminate()
        try:
            proc.wait(timeout=5)
        except subprocess.TimeoutExpired:
            proc.kill()
    except Exception:
        pass

def call_health() -> None:
    with httpx.Client(timeout=5.0) as cx:
        r1 = cx.get(f"{BASE}/healthz")
        r2 = cx.get(f"{BASE}/readyz")
        assert r1.status_code == 200, f"/healthz={r1.status_code}"
        assert r2.status_code == 200, f"/readyz={r2.status_code}"
    _print_ok("Health endpoints OK.")

def call_ask_mock() -> Dict[str,Any]:
    _print_info("Calling POST /ask with provider=mock ...")
    payload = {"query":"ping from quickcheck","provider":"mock","model":"mock-1"}
    headers = {"Authorization": f"Bearer {TOKEN}", "Content-Type":"application/json"}
    with httpx.Client(timeout=20.0) as cx:
        r = cx.post(f"{BASE}/ask", headers=headers, json=payload)
    assert r.status_code == 200, f"/ask status {r.status_code}: {r.text[:200]}"
    # header checks
    x_phase = r.headers.get("X-Perf-Phase")
    x_lat = r.headers.get("X-Latency-MS")
    assert x_phase in {"cold","warm"}, f"X-Perf-Phase invalid: {x_phase}"
    assert x_lat is not None and x_lat.isdigit(), f"X-Latency-MS invalid: {x_lat}"
    body = r.json()
    assert "answer" in body and "context" in body, "Missing fields in /ask response"
    _print_ok(f"/ask OK (phase={x_phase}, latency_ms={x_lat}).")
    return body

def run_orchestrator() -> Tuple[str,Dict[str,Any]]:
    _print_info(f"POST /orchestrator/run_tests (suites: {', '.join(SUITES)}) ...")
    _print_info("This may take 1-2 minutes for test execution...")
    headers = {"Authorization": f"Bearer {TOKEN}", "Content-Type":"application/json"}
    payload = {
        "target_mode":"api",
        "api_base_url": BASE,
        "api_bearer_token": TOKEN,
        "suites": SUITES,
        "options": {"provider":"mock","model":"mock-1"}
    }
    with httpx.Client(timeout=180.0) as cx:
        r = cx.post(f"{BASE}/orchestrator/run_tests", headers=headers, json=payload)
    assert r.status_code == 200, f"/orchestrator/run_tests status {r.status_code}: {r.text[:200]}"
    res = r.json()
    assert "run_id" in res and "artifacts" in res, "Missing run_id/artifacts"
    run_id = res["run_id"]
    _print_ok(f"Run created: {run_id}")
    return run_id, res

def download_artifacts(run_id: str) -> pathlib.Path:
    outdir = pathlib.Path("tmp") / f"quickcheck_{run_id}"
    outdir.mkdir(parents=True, exist_ok=True)
    headers = {"Authorization": f"Bearer {TOKEN}"}
    # JSON
    url_json = f"{BASE}/orchestrator/report/{run_id}.json"
    rj = httpx.get(url_json, headers=headers, timeout=30.0)
    assert rj.status_code == 200, f"GET json status {rj.status_code}"
    p_json = outdir / "report.json"
    p_json.write_bytes(rj.content)
    # XLSX
    url_xlsx = f"{BASE}/orchestrator/report/{run_id}.xlsx"
    rx = httpx.get(url_xlsx, headers=headers, timeout=60.0)
    assert rx.status_code == 200, f"GET xlsx status {rx.status_code}"
    p_xlsx = outdir / "report.xlsx"
    p_xlsx.write_bytes(rx.content)
    _print_ok(f"Artifacts downloaded to {outdir}")
    return outdir

def validate_xlsx(path: pathlib.Path) -> None:
    from openpyxl import load_workbook
    wb = load_workbook(path)
    sheet_names = wb.sheetnames
    for s in REQUIRED_SHEETS:
        assert s in sheet_names, f"Missing sheet '{s}'"
    _print_ok(f"Excel sheets OK: {', '.join(REQUIRED_SHEETS)}")

def maybe_a2a_manifest() -> None:
    try:
        headers = {"Authorization": f"Bearer {TOKEN}"}
        r = httpx.get(f"{BASE}/a2a/manifest", headers=headers, timeout=5.0)
        if r.status_code == 200:
            _print_ok("A2A manifest reachable.")
        else:
            _print_info(f"A2A manifest not available (status {r.status_code}) — skipping.")
    except Exception:
        _print_info("A2A check skipped.")

def main() -> int:
    proc = None
    try:
        proc = start_server()
        wait_ready()
        call_health()
        call_ask_mock()
        run_id, _ = run_orchestrator()
        outdir = download_artifacts(run_id)
        validate_xlsx(outdir / "report.xlsx")
        maybe_a2a_manifest()
        _print_ok("Quick validation PASSED.")
        return 0
    except AssertionError as ae:
        _print_fail(str(ae))
        return 2
    except Exception as e:
        _print_fail(repr(e))
        return 1
    finally:
        stop_server(proc)

if __name__ == "__main__":
    sys.exit(main())
