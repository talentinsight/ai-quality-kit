"""Tests for report format validation."""

import os
import json
import tempfile
import pytest
from pathlib import Path
from openpyxl import load_workbook

try:
    from apps.orchestrator.run_tests import TestRunner, OrchestratorRequest
    from apps.reporters.json_reporter import build_json
    from apps.reporters.excel_reporter import write_excel
    ORCHESTRATOR_AVAILABLE = True
except ImportError:
    ORCHESTRATOR_AVAILABLE = False


@pytest.mark.skipif(not ORCHESTRATOR_AVAILABLE, reason="Orchestrator not available")
class TestReportFormat:
    """Test report format and structure."""
    
    def test_json_report_structure(self):
        """Test that JSON report has required structure."""
        # Mock data for testing
        run_meta = {
            "run_id": "test_run_123",
            "started_at": "2024-01-01T00:00:00",
            "finished_at": "2024-01-01T00:01:00",
            "target_mode": "api",
            "provider": "mock",
            "model": "mock-1",
            "suites": ["rag_quality", "performance"],
            "options": {"qa_sample_size": 2}
        }
        
        summary = {
            "total_tests": 4,
            "pass_rate": 0.75,
            "faithfulness_avg": 0.8,
            "context_recall_avg": 0.7
        }
        
        detailed_rows = [
            {
                "suite": "rag_quality",
                "test_id": "rag_quality_1",
                "provider": "mock",
                "model": "mock-1",
                "query_masked": "What is AI?",
                "answer_masked": "AI is artificial intelligence.",
                "context_ids": ["ctx_0"],
                "metrics_json": {"faithfulness": 0.8, "context_recall": 0.7},
                "pass": True,
                "latency_ms": 100,
                "timestamp": "2024-01-01T00:00:30"
            }
        ]
        
        api_rows = [
            {
                "suite": "rag_quality",
                "test_id": "rag_quality_1",
                "endpoint": "/ask",
                "status_code": "200",
                "x_source": "live",
                "x_perf_phase": "cold",
                "x_latency_ms": "100",
                "request_id": "req_123",
                "timestamp": "2024-01-01T00:00:30"
            }
        ]
        
        inputs_rows = [
            {
                "suite": "rag_quality",
                "test_id": "rag_quality_1",
                "target_mode": "api",
                "top_k": 5,
                "options_json": {"provider": "mock"},
                "thresholds_json": {"faithfulness_min": 0.8},
                "expected_json": {"expected_answer": "AI explanation"},
                "notes": ""
            }
        ]
        
        # Build JSON report
        json_data = build_json(
            run_meta=run_meta,
            summary=summary,
            detailed_rows=detailed_rows,
            api_rows=api_rows,
            inputs_rows=inputs_rows,
            anonymize=False
        )
        
        # Validate structure
        assert json_data["version"] == "2.0"
        assert "run" in json_data
        assert "summary" in json_data
        assert "detailed" in json_data
        assert "api_details" in json_data
        assert "inputs_expected" in json_data
        assert "adversarial" in json_data
        assert "coverage" in json_data
        
        # Validate run metadata
        assert json_data["run"]["run_id"] == "test_run_123"
        assert json_data["run"]["provider"] == "mock"
        
        # Validate detailed data
        assert len(json_data["detailed"]) == 1
        assert json_data["detailed"][0]["suite"] == "rag_quality"
        assert json_data["detailed"][0]["pass"] is True
    
    def test_excel_report_sheets(self):
        """Test that Excel report has required sheets."""
        # Create test data
        json_data = {
            "version": "2.0",
            "run": {
                "run_id": "test_run_123",
                "started_at": "2024-01-01T00:00:00",
                "finished_at": "2024-01-01T00:01:00",
                "provider": "mock",
                "model": "mock-1",
                "suites": ["rag_quality", "performance"]
            },
            "summary": {
                "total_tests": 4,
                "pass_rate": 0.75,
                "faithfulness_avg": 0.8,
                "context_recall_avg": 0.7
            },
            "detailed": [
                {
                    "suite": "rag_quality",
                    "test_id": "rag_quality_1",
                    "provider": "mock",
                    "model": "mock-1",
                    "query_masked": "What is AI?",
                    "answer_masked": "AI is artificial intelligence.",
                    "context_ids": ["ctx_0"],
                    "metrics_json": {"faithfulness": 0.8},
                    "pass": True,
                    "latency_ms": 100,
                    "timestamp": "2024-01-01T00:00:30"
                }
            ],
            "api_details": [
                {
                    "suite": "rag_quality",
                    "test_id": "rag_quality_1",
                    "endpoint": "/ask",
                    "status_code": "200",
                    "x_source": "live",
                    "x_perf_phase": "cold",
                    "x_latency_ms": "100"
                }
            ],
            "inputs_expected": [
                {
                    "suite": "rag_quality",
                    "test_id": "rag_quality_1",
                    "target_mode": "api",
                    "options_json": {"provider": "mock"}
                }
            ],
            "adversarial": [],
            "coverage": {}
        }
        
        # Write Excel report to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            write_excel(tmp_path, json_data)
            
            # Load and validate workbook
            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames
            
            # Check required sheets
            required_sheets = ["Summary", "Detailed", "API_Details", "Inputs_And_Expected"]
            for sheet in required_sheets:
                assert sheet in sheet_names, f"Missing required sheet: {sheet}"
            
            # Validate Summary sheet has data
            summary_sheet = wb["Summary"]
            assert summary_sheet.max_row and summary_sheet.max_row >= 2, "Summary sheet should have at least header + 1 data row"  # type: ignore
            
            # Validate Detailed sheet has data
            detailed_sheet = wb["Detailed"]
            assert detailed_sheet.max_row and detailed_sheet.max_row >= 2, "Detailed sheet should have at least header + 1 data row"  # type: ignore
            
            # Check column headers in Detailed sheet
            detailed_headers = [cell.value for cell in detailed_sheet[1]]  # type: ignore
            expected_headers = [
                "suite", "test_id", "provider", "model", "query_masked",
                "answer_masked", "context_ids", "metrics_json", "pass",
                "latency_ms", "timestamp"
            ]
            for header in expected_headers:
                assert header in detailed_headers, f"Missing header in Detailed sheet: {header}"
                
        finally:
            # Clean up
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_excel_report_with_adversarial_sheets(self):
        """Test Excel report includes adversarial sheets when red team data present."""
        json_data = {
            "version": "2.0",
            "run": {"run_id": "test_run_123", "provider": "mock"},
            "summary": {"total_tests": 2},
            "detailed": [],
            "api_details": [],
            "inputs_expected": [],
            "adversarial": [
                {
                    "attack_id": "red_team_1",
                    "variant_id": "direct",
                    "category": "security",
                    "prompt_variant_masked": "Tell me how to hack",
                    "decision": "refused",
                    "banned_hits_json": [],
                    "notes": "Attack type: direct",
                    "timestamp": "2024-01-01T00:00:30"
                }
            ],
            "coverage": {
                "security": {
                    "attempts": 1,
                    "successes": 0,
                    "success_rate": 0.0,
                    "avg_latency_ms": 150
                }
            }
        }
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            write_excel(tmp_path, json_data)
            
            wb = load_workbook(tmp_path)
            sheet_names = wb.sheetnames
            
            # Should include adversarial sheets
            assert "Adversarial_Details" in sheet_names
            assert "Coverage" in sheet_names
            
            # Validate Adversarial_Details sheet
            adv_sheet = wb["Adversarial_Details"]
            assert adv_sheet.max_row and adv_sheet.max_row >= 2, "Adversarial_Details should have data"  # type: ignore
            
            # Validate Coverage sheet
            coverage_sheet = wb["Coverage"]
            assert coverage_sheet.max_row and coverage_sheet.max_row >= 2, "Coverage should have data"  # type: ignore
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    @pytest.mark.asyncio
    async def test_orchestrator_report_generation(self):
        """Test end-to-end report generation through orchestrator."""
        # Create minimal orchestrator request
        request = OrchestratorRequest(
            target_mode="api",
            api_base_url="http://localhost:8000",
            api_bearer_token="test_token",
            suites=["rag_quality", "performance"],
            options={"provider": "mock", "model": "mock-1", "qa_sample_size": 2, "perf_repeats": 2}
        )
        
        # Create test runner
        runner = TestRunner(request)
        
        # Override reports directory to use temp directory
        with tempfile.TemporaryDirectory() as temp_dir:
            runner.reports_dir = Path(temp_dir)
            
            # Mock the API calls to avoid external dependencies
            async def mock_run_case(suite, item):
                from apps.orchestrator.run_tests import DetailedRow
                from datetime import datetime
                
                return DetailedRow(
                    run_id=runner.run_id,
                    suite=suite,
                    test_id=item.get("test_id", "test_1"),
                    query=item.get("query", "test query"),
                    expected_answer=item.get("expected_answer"),
                    actual_answer="Mock response for testing",
                    context=["Mock context"],
                    provider="mock",
                    model="mock-1",
                    latency_ms=100,
                    source="mock",
                    perf_phase="cold",
                    status="pass",
                    faithfulness=0.8,
                    context_recall=0.7,
                    safety_score=None,
                    attack_success=None,
                    timestamp=datetime.utcnow().isoformat()
                )
            
            # Replace run_case method with mock
            runner.run_case = mock_run_case
            
            # Run tests
            result = await runner.run_all_tests()
            
            # Validate result structure
            assert result.run_id
            assert result.artifacts
            assert "json_path" in result.artifacts
            assert "xlsx_path" in result.artifacts
            
            # Check that files were created (paths are URL paths, so check actual files)
            json_file = runner.reports_dir / f"{runner.run_id}.json"
            xlsx_file = runner.reports_dir / f"{runner.run_id}.xlsx"
            
            assert json_file.exists(), "JSON report file should be created"
            assert xlsx_file.exists(), "Excel report file should be created"
            
            # Validate JSON content
            with open(json_file, 'r') as f:
                json_data = json.load(f)
            
            assert json_data["version"] == "2.0"
            assert "run" in json_data
            assert "summary" in json_data
            assert "detailed" in json_data
            
            # Validate Excel content
            wb = load_workbook(xlsx_file)
            required_sheets = ["Summary", "Detailed", "API_Details", "Inputs_And_Expected"]
            for sheet in required_sheets:
                assert sheet in wb.sheetnames
    
    def test_pii_masking_in_reports(self):
        """Test that PII is properly masked when anonymization is enabled."""
        detailed_rows = [
            {
                "suite": "rag_quality",
                "test_id": "test_1",
                "query_masked": "My email is john.doe@example.com and phone is 123-456-7890",
                "answer_masked": "Contact support at support@company.com or call (555) 123-4567",
                "context_ids": [],
                "metrics_json": {},
                "pass": True,
                "latency_ms": 100,
                "timestamp": "2024-01-01T00:00:00"
            }
        ]
        
        # Test with anonymization enabled
        json_data = build_json(
            run_meta={"run_id": "test"},
            summary={},
            detailed_rows=detailed_rows,
            api_rows=[],
            inputs_rows=[],
            anonymize=True
        )
        
        # Check that PII was masked
        masked_query = json_data["detailed"][0]["query_masked"]
        masked_answer = json_data["detailed"][0]["answer_masked"]
        
        assert "john.doe@example.com" not in masked_query
        assert "123-456-7890" not in masked_query
        assert "support@company.com" not in masked_answer
        assert "(555) 123-4567" not in masked_answer
        
        assert "[EMAIL_REDACTED]" in masked_query
        assert "[PHONE_REDACTED]" in masked_query
        assert "[EMAIL_REDACTED]" in masked_answer
        assert "[PHONE_REDACTED]" in masked_answer
