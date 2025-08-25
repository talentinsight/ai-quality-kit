"""Comprehensive tests for resilience test suite."""

import asyncio
import json
import pytest
from unittest.mock import patch, AsyncMock, MagicMock
from apps.orchestrator.run_tests import TestRunner, OrchestratorRequest
from apps.orchestrator.resilient_client import ResilientClient, CircuitBreaker, ResilienceResult


class TestCircuitBreaker:
    """Test circuit breaker functionality."""
    
    def test_circuit_breaker_initial_state(self):
        """Test circuit breaker starts in closed state."""
        breaker = CircuitBreaker(fail_threshold=3, reset_timeout=10)
        
        assert not breaker.is_open()
        assert breaker.state == "closed"
        assert breaker.failure_count == 0
    
    def test_circuit_breaker_opens_after_threshold(self):
        """Test circuit breaker opens after failure threshold."""
        breaker = CircuitBreaker(fail_threshold=3, reset_timeout=10)
        
        # Record failures
        breaker.record_failure()
        assert not breaker.is_open()
        assert breaker.state == "closed"
        
        breaker.record_failure()
        assert not breaker.is_open()
        assert breaker.state == "closed"
        
        breaker.record_failure()
        assert breaker.is_open()
        assert breaker.state == "open"
    
    def test_circuit_breaker_resets_on_success(self):
        """Test circuit breaker resets failure count on success."""
        breaker = CircuitBreaker(fail_threshold=3, reset_timeout=10)
        
        breaker.record_failure()
        breaker.record_failure()
        assert breaker.failure_count == 2
        
        breaker.record_success()
        assert breaker.failure_count == 0
        assert breaker.state == "closed"
    
    def test_circuit_breaker_half_open_transition(self):
        """Test circuit breaker transitions to half-open after timeout."""
        import time
        
        breaker = CircuitBreaker(fail_threshold=2, reset_timeout=1)
        
        # Trip the breaker
        breaker.record_failure()
        breaker.record_failure()
        assert breaker.is_open()
        
        # Wait for reset timeout
        time.sleep(1.1)
        
        # Should transition to half-open
        assert not breaker.is_open()
        assert breaker.state == "half_open"


class TestResilientClient:
    """Test resilient client functionality."""
    
    @pytest.fixture
    def client(self):
        """Create resilient client."""
        return ResilientClient()
    
    @pytest.mark.asyncio
    async def test_synthetic_call_outcomes(self, client):
        """Test synthetic call generates different outcomes."""
        config = {
            "mode": "synthetic",
            "timeout_ms": 1000,
            "concurrency": 10,
            "queue_depth": 50,
            "circuit": {"fails": 5, "reset_s": 30}
        }
        
        # Run multiple synthetic calls to get distribution
        outcomes = []
        for _ in range(50):
            result = await client._make_synthetic_call("test_req", config)
            outcomes.append(result.outcome)
        
        # Should have variety of outcomes
        outcome_set = set(outcomes)
        assert "success" in outcome_set
        assert len(outcome_set) >= 2  # At least success and one failure type
        
        # All results should have proper structure
        for _ in range(3):
            result = await client._make_synthetic_call("test_req", config)
            assert result.request_id == "test_req"
            assert result.mode == "synthetic"
            assert result.latency_ms >= 0
            assert result.attempts == 1
            assert result.outcome in ["success", "timeout", "upstream_5xx", "upstream_429"]
    
    @pytest.mark.asyncio
    async def test_circuit_breaker_integration(self, client):
        """Test circuit breaker integration with resilient client."""
        config = {
            "mode": "synthetic",
            "timeout_ms": 1000,
            "concurrency": 10,
            "queue_depth": 50,
            "circuit": {"fails": 2, "reset_s": 30}  # Low threshold for testing
        }
        
        # Force circuit breaker to open by recording failures
        circuit_breaker = client._get_circuit_breaker("test:mock", config)
        circuit_breaker.record_failure()
        circuit_breaker.record_failure()
        
        # Next call should return circuit_open
        result = await client.call_with_resilience("test query", "test", "mock", config)
        assert result.outcome == "circuit_open"
        assert result.error_class == "CircuitBreakerOpenError"
        assert result.latency_ms == 0
        assert result.attempts == 0
    
    @pytest.mark.asyncio
    async def test_concurrency_control(self, client):
        """Test concurrency control with semaphore."""
        config = {
            "mode": "synthetic",
            "timeout_ms": 100,
            "concurrency": 2,  # Low concurrency for testing
            "queue_depth": 50,
            "circuit": {"fails": 5, "reset_s": 30}
        }
        
        # Create multiple concurrent calls
        tasks = []
        for i in range(5):
            task = client.call_with_resilience(f"query_{i}", "test", "mock", config)
            tasks.append(task)
        
        results = await asyncio.gather(*tasks)
        
        # All should complete successfully (synthetic mode)
        assert len(results) == 5
        for result in results:
            assert result.outcome in ["success", "timeout", "upstream_5xx", "upstream_429"]


class TestResilienceOrchestrator:
    """Test resilience suite integration with orchestrator."""
    
    @pytest.fixture
    def orchestrator_request(self):
        """Create orchestrator request with resilience suite."""
        return OrchestratorRequest(
            target_mode="api",
            suites=["resilience"],
            options={
                "provider": "mock",
                "model": "mock-1",
                "resilience": {
                    "mode": "synthetic",
                    "samples": 3,
                    "timeout_ms": 1000,
                    "retries": 0,
                    "concurrency": 5,
                    "queue_depth": 20,
                    "circuit": {"fails": 3, "reset_s": 30}
                }
            }
        )
    
    def test_load_resilience_tests(self, orchestrator_request):
        """Test loading resilience tests from request."""
        runner = TestRunner(orchestrator_request)
        suite_data = runner.load_suites()
        
        assert "resilience" in suite_data
        resilience_tests = suite_data["resilience"]
        assert len(resilience_tests) == 48  # Full resilience catalog
        
        for test in resilience_tests:
            assert test["test_type"] == "resilience"
            assert test["category"] == "robustness"
            assert "test_id" in test
            assert "query" in test
            assert "resilience_config" in test
            
            config = test["resilience_config"]
            assert config["mode"] == "synthetic"
            assert config["timeout_ms"] == 1000
            assert config["retries"] == 0
            assert config["concurrency"] == 5
    
    @pytest.mark.asyncio
    async def test_run_resilience_case(self, orchestrator_request):
        """Test running a single resilience case."""
        runner = TestRunner(orchestrator_request)
        
        test_item = {
            "test_id": "resilience_probe_1",
            "query": "What is the AI Quality Kit?",
            "test_type": "resilience",
            "resilience_config": {
                "mode": "synthetic",
                "timeout_ms": 1000,
                "retries": 0,
                "concurrency": 5,
                "queue_depth": 20,
                "circuit": {"fails": 3, "reset_s": 30}
            }
        }
        
        result = await runner._run_resilience_case(test_item, "mock", "mock-1")
        
        # Check result structure
        assert "answer" in result
        assert "resilience_outcome" in result
        assert "resilience_latency_ms" in result
        assert "resilience_attempts" in result
        assert result["source"] == "resilience_test"
        assert result["perf_phase"] == "resilience"
        
        # Check resilience details were stored
        assert len(runner.resilience_details) == 1
        detail = runner.resilience_details[0]
        assert detail["run_id"] == runner.run_id
        assert detail["provider"] == "mock"
        assert detail["model"] == "mock-1"
        assert detail["outcome"] in ["success", "timeout", "upstream_5xx", "upstream_429"]
        assert detail["mode"] == "synthetic"
    
    @pytest.mark.asyncio
    async def test_resilience_summary_calculation(self, orchestrator_request):
        """Test resilience summary calculation."""
        runner = TestRunner(orchestrator_request)
        
        # First load and run some tests to populate detailed_rows
        suite_data = runner.load_suites()
        assert "resilience" in suite_data
        
        # Add some dummy detailed rows so summary generation works
        from apps.orchestrator.run_tests import DetailedRow
        dummy_row = DetailedRow(
            run_id=runner.run_id,
            suite="resilience",
            test_id="test_1",
            query="test query",
            expected_answer=None,
            actual_answer="test answer",
            context=[],
            provider="mock",
            model="mock-1",
            latency_ms=100,
            source="test",
            perf_phase="test",
            status="pass",
            faithfulness=None,
            context_recall=None,
            safety_score=None,
            attack_success=None,
            timestamp="2024-01-01T00:00:00Z"
        )
        runner.detailed_rows.append(dummy_row)
        
        # Simulate some resilience details
        runner.resilience_details = [
            {
                "run_id": runner.run_id,
                "timestamp": 1234567890,
                "provider": "mock",
                "model": "mock-1",
                "request_id": "req_1",
                "outcome": "success",
                "attempts": 1,
                "latency_ms": 100,
                "error_class": None,
                "mode": "synthetic"
            },
            {
                "run_id": runner.run_id,
                "timestamp": 1234567891,
                "provider": "mock",
                "model": "mock-1", 
                "request_id": "req_2",
                "outcome": "timeout",
                "attempts": 1,
                "latency_ms": 1000,
                "error_class": "TimeoutError",
                "mode": "synthetic"
            },
            {
                "run_id": runner.run_id,
                "timestamp": 1234567892,
                "provider": "mock",
                "model": "mock-1",
                "request_id": "req_3",
                "outcome": "success",
                "attempts": 1,
                "latency_ms": 200,
                "error_class": None,
                "mode": "synthetic"
            }
        ]
        
        summary = runner._generate_summary()
        
        assert "resilience" in summary
        resilience_summary = summary["resilience"]
        
        assert resilience_summary["samples"] == 3
        assert resilience_summary["success_rate"] == 2/3  # 2 out of 3 successful
        assert resilience_summary["timeouts"] == 1
        assert resilience_summary["upstream_5xx"] == 0
        assert resilience_summary["upstream_429"] == 0
        assert resilience_summary["circuit_open_events"] == 0
        
        # Should have percentiles from successful attempts only
        assert "p50_ms" in resilience_summary
        assert "p95_ms" in resilience_summary
        assert resilience_summary["p50_ms"] in [100, 200]  # Median of [100, 200]


class TestBackwardCompatibility:
    """Test backward compatibility for gibberish -> resilience alias."""
    
    @pytest.fixture
    def gibberish_request(self):
        """Create orchestrator request with gibberish suite."""
        return OrchestratorRequest(
            target_mode="api",
            suites=["gibberish"],
            options={"provider": "mock", "model": "mock-1"}
        )
    
    def test_gibberish_alias_mapping(self, gibberish_request):
        """Test gibberish suite gets mapped to resilience."""
        runner = TestRunner(gibberish_request)
        suite_data = runner.load_suites()
        
        # Should have resilience tests, not gibberish
        assert "resilience" in suite_data
        assert "gibberish" not in suite_data
        
        # Should track deprecated suite
        assert "gibberish" in runner.deprecated_suites
    
    def test_deprecation_note_in_summary(self, gibberish_request):
        """Test deprecation note appears in summary."""
        runner = TestRunner(gibberish_request)
        suite_data = runner.load_suites()  # Triggers the alias mapping
        
        # Ensure we have resilience tests loaded
        assert "resilience" in suite_data
        assert len(runner.deprecated_suites) > 0
        
        # Add dummy detailed row for summary generation
        from apps.orchestrator.run_tests import DetailedRow
        dummy_row = DetailedRow(
            run_id=runner.run_id,
            suite="resilience",
            test_id="test_1",
            query="test query",
            expected_answer=None,
            actual_answer="test answer",
            context=[],
            provider="mock",
            model="mock-1",
            latency_ms=100,
            source="test",
            perf_phase="test",
            status="pass",
            faithfulness=None,
            context_recall=None,
            safety_score=None,
            attack_success=None,
            timestamp="2024-01-01T00:00:00Z"
        )
        runner.detailed_rows.append(dummy_row)
        
        summary = runner._generate_summary()
        
        assert "_deprecated_note" in summary
        assert "gibberish → resilience" in summary["_deprecated_note"]


class TestResilienceReporting:
    """Test resilience data in reports."""
    
    @pytest.mark.asyncio 
    async def test_json_report_resilience_section(self):
        """Test JSON report includes resilience section."""
        from apps.reporters.json_reporter import build_json
        
        # Mock resilience details
        resilience_details = [
            {
                "run_id": "test_run",
                "timestamp": 1234567890,
                "provider": "mock",
                "model": "mock-1",
                "request_id": "req_1",
                "outcome": "success",
                "attempts": 1,
                "latency_ms": 150,
                "error_class": None,
                "mode": "passive"
            }
        ]
        
        # Mock summary with resilience data
        summary = {
            "resilience": {
                "samples": 1,
                "success_rate": 1.0,
                "timeouts": 0,
                "upstream_5xx": 0,
                "upstream_429": 0,
                "circuit_open_events": 0,
                "p50_ms": 150,
                "p95_ms": 150
            }
        }
        
        run_meta = {"run_id": "test_run", "provider": "mock", "model": "mock-1"}
        
        report = build_json(
            run_meta=run_meta,
            summary=summary,
            detailed_rows=[],
            api_rows=[],
            inputs_rows=[],
            resilience_details=resilience_details,
            anonymize=False
        )
        
        # Should have resilience section
        assert "resilience" in report
        assert "summary" in report["resilience"]
        assert "details" in report["resilience"]
        
        # Check summary
        resilience_summary = report["resilience"]["summary"]
        assert resilience_summary["samples"] == 1
        assert resilience_summary["success_rate"] == 1.0
        
        # Check details
        details = report["resilience"]["details"]
        assert len(details) == 1
        assert details[0]["outcome"] == "success"
        assert details[0]["latency_ms"] == 150
    
    def test_excel_report_resilience_sheet(self):
        """Test Excel report includes Resilience_Details sheet."""
        import tempfile
        import os
        from openpyxl import load_workbook
        from apps.reporters.excel_reporter import write_excel
        
        # Mock data with resilience section
        data = {
            "version": "2.0",
            "run": {"run_id": "test_run", "provider": "mock", "model": "mock-1"},
            "summary": {
                "overall": {"total_tests": 1, "pass_rate": 1.0}
            },
            "detailed": [],
            "api_details": [],
            "inputs_expected": [],
            "adversarial_details": [],
            "coverage": {},
            "resilience": {
                "summary": {
                    "samples": 2,
                    "success_rate": 0.5,
                    "timeouts": 1,
                    "upstream_5xx": 0,
                    "upstream_429": 0,
                    "circuit_open_events": 0,
                    "p50_ms": 100,
                    "p95_ms": 200
                },
                "details": [
                    {
                        "run_id": "test_run",
                        "timestamp": 1234567890,
                        "provider": "mock",
                        "model": "mock-1",
                        "request_id": "req_1",
                        "outcome": "success",
                        "attempts": 1,
                        "latency_ms": 100,
                        "error_class": None,
                        "mode": "passive"
                    },
                    {
                        "run_id": "test_run", 
                        "timestamp": 1234567891,
                        "provider": "mock",
                        "model": "mock-1",
                        "request_id": "req_2",
                        "outcome": "timeout",
                        "attempts": 1,
                        "latency_ms": 5000,
                        "error_class": "TimeoutError",
                        "mode": "passive"
                    }
                ]
            }
        }
        
        # Write to temporary file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            tmp_path = tmp_file.name
        
        try:
            write_excel(tmp_path, data)
            
            # Load and verify
            wb = load_workbook(tmp_path)
            
            # Should have Resilience_Details sheet
            assert "Resilience_Details" in wb.sheetnames
            
            sheet = wb["Resilience_Details"]
            
            # Check headers (row 1)
            expected_headers = [
                "run_id", "timestamp", "provider", "model", "request_id",
                "outcome", "attempts", "latency_ms", "error_class", "mode"
            ]
            
            for col, expected_header in enumerate(expected_headers, 1):
                actual_header = sheet.cell(row=1, column=col).value  # type: ignore
                assert actual_header == expected_header
            
            # Check data rows
            assert sheet.cell(row=2, column=6).value == "success"  # type: ignore  # outcome
            assert sheet.cell(row=2, column=8).value == 100        # type: ignore  # latency_ms
            assert sheet.cell(row=3, column=6).value == "timeout"  # type: ignore  # outcome
            assert sheet.cell(row=3, column=8).value == 5000       # type: ignore  # latency_ms
            
        finally:
            os.unlink(tmp_path)


if __name__ == "__main__":
    pytest.main([__file__])
