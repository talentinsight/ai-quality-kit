"""Unit tests for RAG-specific Excel reporter functionality."""

import pytest
import tempfile
from pathlib import Path
from openpyxl import load_workbook

from apps.reporters.excel_reporter import write_excel


class TestRAGExcelReporter:
    """Test RAG-specific Excel reporting functionality."""
    
    def test_summary_sheet_rag_fields(self):
        """Test that Summary sheet includes RAG-specific fields."""
        # Create test data with RAG results
        test_data = {
            "run": {
                "run_id": "test_run_123",
                "started_at": "2024-01-01T10:00:00Z",
                "finished_at": "2024-01-01T10:05:00Z",
                "target_mode": "api",
                "ground_truth": "available",
                "provider": "openai",
                "model": "gpt-4",
                "suites": ["rag_quality"],
                "gate": True,
                "elapsed_ms": 5000
            },
            "summary": {
                "overall": {
                    "total_tests": 10,
                    "passed": 8,
                    "pass_rate": 0.8
                },
                "rag_quality": {
                    "total": 5,
                    "passed": 4,
                    "pass_rate": 0.8,
                    "gate": True,
                    "avg_faithfulness": 0.85,
                    "avg_context_recall": 0.82,
                    "avg_answer_relevancy": 0.78
                }
            },
            "detailed": [],
            "api_details": [],
            "inputs_expected": []
        }
        
        # Write Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            write_excel(tmp_file.name, test_data)
            
            # Load and verify
            wb = load_workbook(tmp_file.name)
            
            # Check Summary sheet exists
            assert "Summary" in wb.sheetnames
            ws = wb["Summary"]
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Verify RAG-specific headers are present
            assert "target_mode" in headers
            assert "ground_truth" in headers
            assert "gate" in headers
            assert "elapsed_ms" in headers
            
            # Get data from second row
            data_row = [cell.value for cell in ws[2]]
            data_dict = dict(zip(headers, data_row))
            
            # Verify RAG-specific data
            assert data_dict["target_mode"] == "api"
            assert data_dict["ground_truth"] == "available"
            assert data_dict["gate"] == True
            assert data_dict["elapsed_ms"] == 5000
            
            # Clean up
            Path(tmp_file.name).unlink()
    
    def test_detailed_sheet_rag_columns(self):
        """Test that Detailed sheet includes RAG-specific columns."""
        # Create test data with detailed RAG results
        test_data = {
            "run": {
                "run_id": "test_run_123",
                "started_at": "2024-01-01T10:00:00Z",
                "finished_at": "2024-01-01T10:05:00Z"
            },
            "summary": {"overall": {"total_tests": 1}},
            "detailed": [
                {
                    "suite": "rag_quality",
                    "case_id": "q1",
                    "provider": "openai",
                    "model": "gpt-4",
                    "question": "What is the capital of France?",
                    "predicted": "The capital of France is Paris.",
                    "expected": "Paris",
                    "retrieved_count": 3,
                    "faithfulness": 0.85,
                    "context_recall": 0.82,
                    "answer_relevancy": 0.78,
                    "context_precision": 0.75,
                    "answer_correctness": 0.80,
                    "answer_similarity": 0.77,
                    "pass_fail_reason": "All metrics above threshold",
                    "latency_ms": 1200,
                    "timestamp": "2024-01-01T10:01:00Z"
                }
            ],
            "api_details": [],
            "inputs_expected": []
        }
        
        # Write Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            write_excel(tmp_file.name, test_data)
            
            # Load and verify
            wb = load_workbook(tmp_file.name)
            
            # Check Detailed sheet exists
            assert "Detailed" in wb.sheetnames
            ws = wb["Detailed"]
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Verify RAG-specific headers are present
            expected_headers = [
                "suite", "case_id", "provider", "model", "question", "predicted",
                "expected", "retrieved_count", "recall_at_k", "mrr_at_k", "ndcg_at_k",
                "faithfulness", "context_recall", "answer_relevancy", "context_precision",
                "answer_correctness", "answer_similarity", "perturbations_applied",
                "pass_fail_reason", "latency_ms", "timestamp"
            ]
            
            for header in expected_headers:
                assert header in headers
            
            # Get data from second row
            data_row = [cell.value for cell in ws[2]]
            data_dict = dict(zip(headers, data_row))
            
            # Verify RAG-specific data
            assert data_dict["case_id"] == "q1"
            assert data_dict["question"] == "What is the capital of France?"
            assert data_dict["predicted"] == "The capital of France is Paris."
            assert data_dict["expected"] == "Paris"
            assert data_dict["retrieved_count"] == 3
            assert data_dict["faithfulness"] == 0.85
            assert data_dict["context_recall"] == 0.82
            
            # Clean up
            Path(tmp_file.name).unlink()
    
    def test_inputs_expected_validation_columns(self):
        """Test that Inputs_And_Expected sheet includes validation columns."""
        # Create test data with validation results
        test_data = {
            "run": {
                "run_id": "test_run_123",
                "started_at": "2024-01-01T10:00:00Z",
                "finished_at": "2024-01-01T10:05:00Z"
            },
            "summary": {"overall": {"total_tests": 1}},
            "detailed": [],
            "api_details": [],
            "inputs_expected": [
                {
                    "suite": "rag_quality",
                    "test_id": "rag_evaluation",
                    "target_mode": "api",
                    "top_k": 3,
                    "options_json": {"temperature": 0.0},
                    "thresholds_json": {"faithfulness": 0.8},
                    "expected_json": {"metrics": ["faithfulness", "context_recall"]},
                    "notes": "RAG quality evaluation with ground truth",
                    "validation_status": "valid",
                    "valid_count": 10,
                    "invalid_count": 0,
                    "duplicate_count": 1,
                    "easy_ratio": 0.1
                }
            ]
        }
        
        # Write Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            write_excel(tmp_file.name, test_data)
            
            # Load and verify
            wb = load_workbook(tmp_file.name)
            
            # Check Inputs_And_Expected sheet exists
            assert "Inputs_And_Expected" in wb.sheetnames
            ws = wb["Inputs_And_Expected"]
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Verify validation-specific headers are present
            validation_headers = [
                "validation_status", "valid_count", "invalid_count", 
                "duplicate_count", "easy_ratio"
            ]
            
            for header in validation_headers:
                assert header in headers
            
            # Get data from second row
            data_row = [cell.value for cell in ws[2]]
            data_dict = dict(zip(headers, data_row))
            
            # Verify validation data
            assert data_dict["validation_status"] == "valid"
            assert data_dict["valid_count"] == 10
            assert data_dict["invalid_count"] == 0
            assert data_dict["duplicate_count"] == 1
            assert data_dict["easy_ratio"] == 0.1
            
            # Clean up
            Path(tmp_file.name).unlink()
    
    def test_coverage_sheet_rag_fields(self):
        """Test that Coverage sheet includes RAG-specific fields."""
        # Create test data with coverage information
        test_data = {
            "run": {
                "run_id": "test_run_123",
                "started_at": "2024-01-01T10:00:00Z",
                "finished_at": "2024-01-01T10:05:00Z"
            },
            "summary": {"overall": {"total_tests": 1}},
            "detailed": [],
            "api_details": [],
            "inputs_expected": [],
            "coverage": {
                "modules": [
                    {
                        "module": "apps.orchestrator.rag_runner",
                        "stmts": 100,
                        "miss": 20,
                        "branch": 50,
                        "brpart": 10,
                        "cover_percent": 80.0,
                        "total_lines": 120,
                        "total_cases": 15,
                        "filtered_cases": 12,
                        "invalid_cases": 2,
                        "filter_reason": "Missing context references"
                    }
                ]
            }
        }
        
        # Write Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            write_excel(tmp_file.name, test_data)
            
            # Load and verify
            wb = load_workbook(tmp_file.name)
            
            # Check Coverage sheet exists
            assert "Coverage" in wb.sheetnames
            ws = wb["Coverage"]
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Verify RAG-specific coverage headers are present
            rag_headers = ["total_cases", "filtered_cases", "invalid_cases", "filter_reason"]
            
            for header in rag_headers:
                assert header in headers
            
            # Get data from second row
            data_row = [cell.value for cell in ws[2]]
            data_dict = dict(zip(headers, data_row))
            
            # Verify RAG coverage data
            assert data_dict["total_cases"] == 15
            assert data_dict["filtered_cases"] == 12
            assert data_dict["invalid_cases"] == 2
            assert data_dict["filter_reason"] == "Missing context references"
            
            # Clean up
            Path(tmp_file.name).unlink()
    
    def test_all_sheets_created(self):
        """Test that all required sheets are created for RAG reports."""
        # Create comprehensive test data
        test_data = {
            "run": {
                "run_id": "test_run_123",
                "started_at": "2024-01-01T10:00:00Z",
                "finished_at": "2024-01-01T10:05:00Z",
                "target_mode": "api",
                "ground_truth": "available"
            },
            "summary": {
                "overall": {"total_tests": 5, "passed": 4, "pass_rate": 0.8}
            },
            "detailed": [{"suite": "rag_quality", "test_id": "test1"}],
            "api_details": [{"endpoint": "/test", "status": 200}],
            "inputs_expected": [{"suite": "rag_quality", "test_id": "config1"}],
            "coverage": {
                "modules": [{"module": "test_module", "stmts": 10, "miss": 2}]
            }
        }
        
        # Write Excel file
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
            write_excel(tmp_file.name, test_data)
            
            # Load and verify all sheets exist
            wb = load_workbook(tmp_file.name)
            
            required_sheets = ["Summary", "Detailed", "API_Details", "Inputs_And_Expected", "Coverage"]
            
            for sheet_name in required_sheets:
                assert sheet_name in wb.sheetnames
            
            # Clean up
            Path(tmp_file.name).unlink()
